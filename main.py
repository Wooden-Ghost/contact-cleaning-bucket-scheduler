from __future__ import annotations

import copy
import json
from collections import deque
import hashlib
import os
import random
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from PySide6.QtCore import QTimer, Qt, Signal
    from PySide6.QtGui import QGuiApplication, QTextOption
    from PySide6.QtWidgets import (
        QApplication,
        QButtonGroup,
        QCheckBox,
        QDialog,
        QFileDialog,
        QFrame,
        QGridLayout,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPushButton,
        QPlainTextEdit,
        QRadioButton,
        QSizePolicy,
        QScrollArea,
        QStatusBar,
        QTabWidget,
        QToolButton,
        QVBoxLayout,
        QWidget,
    )
except ImportError as exc:
    print("当前环境缺少 PySide6，请先安装 PySide6 后再运行。")
    print(exc)
    sys.exit(1)


APP_VERSION = "v12.0.4"
APP_NAME = f"联系人清洗 · 周分桶 {APP_VERSION}"
APP_DIR = Path.home() / ".contact_template_flow_v12_0_0"
LEGACY_APP_DIR = Path.home() / ".contact_template_flow_v11_8_2"
CONFIG_DIR = APP_DIR / "configs"
STATE_PATH = APP_DIR / "state.json"
DEFAULT_CONFIG_PATH = CONFIG_DIR / "default_config.json"
DEFAULT_EXPORT_DIR = str(Path.home() / "Desktop")
DEFAULT_SOURCE_DIR = ""
INDEX_SUBDIR_NAME = "索引文件"
INDEX_JSON_NAME = "【勿删】处理索引.json"
INDEX_XLSX_NAME = "【勿删】处理索引.xlsx"
MAX_EXPORT_COL_WIDTH = 33
TEMPLATE2_TIME_RE = re.compile(r"^(\d{4})/(\d{2})/(\d{2})\|(\d{2}):(00|10|20|30|40|50)$")
GENERATED_ROW_ID_RE = re.compile(r"^[A-Z0-9]{2}(\d{4}[AP]\d{3})$", re.I)
DEFAULT_BUCKET_CONFIG_JSON = json.dumps(
    {
        "Mon": 50,
        "Tue": 50,
        "Wed": 50,
        "Thu": 50,
        "Fri": 50,
    },
    ensure_ascii=False,
    indent=2,
)
DEFAULT_MORNING_PROMPT = """你是冷邮件“模板2计划营销时间”生成助手。

你的任务是：
根据联系人公司的公开信息与地址线索，推断联系人更可能所在的地区，并找到一个合适的联系人当地正常工作日；固定使用该当地工作日上午基准时间 10:10，再将该当地时间换算为中国时间（Asia/Shanghai），输出模板2结果。

你会收到这些信息：
- current_utc
- target_cn_date
- allowed_end_cn_date
- 多行联系人数据，每行格式为：row_id|company_name|address_text

你的输出规则必须绝对严格遵守：

一、每个 row_id 只输出一行结果，格式只允许以下两种之一：
1. row_id|YYYY/MM/DD|HH:MM
2. row_id|NO_VALID_TIME

二、禁止输出任何其他内容：
- 不要解释
- 不要理由
- 不要 JSON
- 不要 markdown
- 不要表头
- 不要空行
- 不要额外符号
- 不要补充说明

三、判断规则：
1. 先根据地址与公司信息判断联系人更可能所在地区。
2. 当地基准时间固定为上午 10:10，不要改成其他时间。
3. 最终输出的 YYYY/MM/DD|HH:MM 必须是中国时间（Asia/Shanghai）。
4. 输出的中国时间必须晚于 current_utc。
5. 优先选择一个联系人当地正常工作日，使其对应的中国时间日期等于 target_cn_date。
6. 如果无法得到满足 target_cn_date 的有效中国时间，则允许顺延到最近可行的未来中国工作日。
7. 顺延结果不得晚于 allowed_end_cn_date。
8. 联系人当地所选日期必须是正常工作日，不要落在周六、周日或其他节假日。
9. 若地址不完整、模糊或不足以稳定判断地区，则结合公司名推断公司总部所在地，并以公司总部所在地为准。
10. 只有在无法较可靠判断联系人所在地区或公司总部所在地，或者无法据此得到有效未来时间时，才输出：row_id|NO_VALID_TIME。
11. 如果 target_cn_date 可行，优先输出 target_cn_date 对应结果；只有 target_cn_date 不可行时，才允许顺延。
12. 不要为了凑 target_cn_date 而输出早于 current_utc 的中国时间。
13. 日期格式必须使用：YYYY/MM/DD|HH:MM。

四、稳定性要求：
1. 输出行数必须与输入行数一致。
2. 每个 row_id 只出现一次。
3. 不要遗漏任何 row_id。
4. 不要改写 row_id。"""
DEFAULT_AFTERNOON_PROMPT = """你是冷邮件“模板2计划营销时间”生成助手。

你的任务是：
根据联系人公司的公开信息与地址线索，推断联系人更可能所在的地区，并找到一个合适的联系人当地正常工作日；固定使用该当地工作日下午基准时间 15:20，再将该当地时间换算为中国时间（Asia/Shanghai），输出模板2结果。

你会收到这些信息：
- current_utc
- target_cn_date
- allowed_end_cn_date
- 多行联系人数据，每行格式为：row_id|company_name|address_text

你的输出规则必须绝对严格遵守：

一、每个 row_id 只输出一行结果，格式只允许以下两种之一：
1. row_id|YYYY/MM/DD|HH:MM
2. row_id|NO_VALID_TIME

二、禁止输出任何其他内容：
- 不要解释
- 不要理由
- 不要 JSON
- 不要 markdown
- 不要表头
- 不要空行
- 不要额外符号
- 不要补充说明

三、判断规则：
1. 先根据地址与公司信息判断联系人更可能所在地区。
2. 当地基准时间固定为下午 15:20，不要改成其他时间。
3. 最终输出的 YYYY/MM/DD|HH:MM 必须是中国时间（Asia/Shanghai）。
4. 输出的中国时间必须晚于 current_utc。
5. 优先选择一个联系人当地正常工作日，使其对应的中国时间日期等于 target_cn_date。
6. 如果无法得到满足 target_cn_date 的有效中国时间，则允许顺延到最近可行的未来中国工作日。
7. 顺延结果不得晚于 allowed_end_cn_date。
8. 联系人当地所选日期必须是正常工作日，不要落在周六、周日或其他节假日。
9. 若地址不完整、模糊或不足以稳定判断地区，则结合公司名推断公司总部所在地，并以公司总部所在地为准。
10. 只有在无法较可靠判断联系人所在地区或公司总部所在地，或者无法据此得到有效未来时间时，才输出：row_id|NO_VALID_TIME。
11. 如果 target_cn_date 可行，优先输出 target_cn_date 对应结果；只有 target_cn_date 不可行时，才允许顺延。
12. 不要为了凑 target_cn_date 而输出早于 current_utc 的中国时间。
13. 日期格式必须使用：YYYY/MM/DD|HH:MM。

四、稳定性要求：
1. 输出行数必须与输入行数一致。
2. 每个 row_id 只出现一次。
3. 不要遗漏任何 row_id。
4. 不要改写 row_id。"""
PROGRAM_INFO_TEXT = """内部更新：v13.7.0
对外窗口名与 RPA 路径保持 12.0.4 不变。

定位

清洗程序负责：
- 清洗联系人
- 单公司/多公司导入
- 按导入顺序处理多公司
- 按周桶分配
- 生成模板1
- 生成模板2输入稿
- 生成模板3
- 输出顺延待处理表
- 输出异常清单
- 生成可直接粘贴到 Excel 的公司日志与联系人日志文本

清洗程序不负责：
- 直接写入正式公司日志
- 直接写入正式联系人日志
- 按天日志目录管理
- 正式状态台账

RPA 负责：
- 读取日志和总表
- 决定做哪些公司
- 决定走批量还是单独
- 复制程序生成的日志文本回正式表
- 写状态、处理时间、完成时间
- 维护总表

说明：
- 首页只保留流程区，不显示完整提示词正文。
- 配置编辑已收回到程序内设置页完成。
- AI 只按当前桶生成模板2。
- 分桶启用基准前 120 分钟截止：AM 截止 08:10，PM 截止 13:20。
- 新建模板2输入稿前会先检查分桶窗口是否已过期，避免继续使用失效桶。
- 分桶单选按钮改为固定槽位编号，便于八爪鱼稳定定位。
- 分桶映射顺序已改为第一列从上到下排完，再进入下一列。
- 桶按钮状态直接写入 Name 和 AutomationId 后缀：有桶且有人=enabled；0/容量或无桶=disabled，便于八爪鱼直接按路径判断。
- RPA 元素路径冻结：后续版本默认不得修改 bucketColumnsHost / bucketColumn__{x} / bucketPickRadio__{x}_{y}__{state} / bucket_slot_{x}_{y}_{state}。
- 导出结果改为按批次维护同一份结果包：同一批覆盖写回同一文件，不同批次新建新文件。
- 公司若出现未分配清单，公司日志状态输出为“有顺延”。
- 计划营销时间必须晚于本次生成模板2输入稿的冻结时间。
- “复制公司日志”“复制联系人日志”输出为制表符分列文本，可直接粘贴到 Excel。
- 启动时先读取【数据源目录】与【索引文件】，若存在已到可处理时间的顺延文件，会先弹出勾选清单。
- 导出结果文件名改为：yyyy-MM-dd xx家共xx人.xlsx。
- 导出目录下新增：索引文件\【勿删】处理索引.json / 【勿删】处理索引.xlsx。
- 顺延待处理文件改为复制原表并删减数据行，尽量保留原表样式。
- 批次总览改为：文件名 / 导出时状态 / 此次入桶数 / 总人数。
- 索引文件保留导出时状态与最终状态两层字段，便于 RPA 后续回写。
- 导出失败提示改为按步骤说明：写结果包 / 顺延文件 / 归档 / 索引。
- 程序重开后默认不自动载入旧批次，桶保持清空。
- 设置页周桶配置改为按“单账号日容量”填写；当前按 6 个账号汇总入桶，总日容量=设置值×6，程序内部再自动拆成 AM / PM。
- 联系人日志输出调整为：姓名后紧跟职业。
- 导入整批文件时即按全局规则预分配营销账号；复制联系人日志时末列输出固定的“营销账号标识”，并严格按账号块连续导出，便于批量处理。
- 修复营销账号分配 finalize_assigned_entries 的边界循环风险：改为有限次数查找，不再使用可能误判为卡死的 while 指针推进。
- 修复营销账号分配异常时界面可能看似失去响应的问题：分配失败会弹窗并安全回退界面状态，不再让主界面停在半加载状态。"""
WEEKDAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri"]
WEEKDAY_LABELS = {"Mon": "周一", "Tue": "周二", "Wed": "周三", "Thu": "周四", "Fri": "周五"}
MODE_LABELS = {"AM": "上午", "PM": "下午"}
MODE_BASE_TIME = {"AM": "10:10", "PM": "15:20"}
MODE_CUTOFF_TIME = {"AM": "08:10", "PM": "13:20"}
ACCOUNT_MARKERS = ["Export", "Calvin", "Contact", "JohnWu", "sales", "GavinZhao"]
ACCOUNT_COUNT = len(ACCOUNT_MARKERS)
ACCOUNT_INDEX_BY_NAME = {name: idx for idx, name in enumerate(ACCOUNT_MARKERS)}
ACCOUNT_DISPLAY_ORDER = list(ACCOUNT_MARKERS)
ACCOUNT_DISPLAY_MAP = {name: idx for idx, name in enumerate(ACCOUNT_DISPLAY_ORDER)}
ACCOUNT_SELECTION_PRIORITY = ["JohnWu", "Calvin", "Export", "Contact", "sales", "GavinZhao"]
ACCOUNT_SELECTION_INDICES = [ACCOUNT_INDEX_BY_NAME[name] for name in ACCOUNT_SELECTION_PRIORITY]
ACCOUNT_WEIGHT_MAP = {
    "Export": 1,
    "Calvin": 2,
    "Contact": 1,
    "JohnWu": 3,
    "sales": 1,
    "GavinZhao": 1,
}
CN_TZ = timezone(timedelta(hours=8))
FIXED_BUCKET_COLUMNS = 4
FIXED_BUCKET_ROWS = 5
FIXED_BUCKET_SLOT_COUNT = FIXED_BUCKET_COLUMNS * FIXED_BUCKET_ROWS
# ==============================
# RPA 路径冻结说明（强约束）
# 1. 自 v12.0.2 起，八爪鱼 RPA 使用的元素路径接口已冻结。
# 2. 以下定位相关字段，除非用户明确要求“重做定位规则”，否则禁止修改：
#    - bucketColumnsHost
#    - bucketColumn__{x}
#    - bucketPickRadio__{x}_{y}__{state}
#    - bucket_slot_{x}_{y}_{state}
#    - RPA_BUCKET_NAME_TEMPLATE
#    - RPA_BUCKET_AUTOMATION_ID_TEMPLATE
#    - RPA_BUCKET_XPATH_TEMPLATE
# 3. 后续版本允许修改：导出逻辑、日志逻辑、批次逻辑、解析逻辑；
#    但不允许擅自修改上述元素路径接口，否则会导致八爪鱼定位失效。
# 4. 本文件在 v12.0.4 中继续沿用 v12.0.2 的全部元素路径命名，不得擅改。
# ==============================
RPA_BUCKET_NAME_TEMPLATE = "bucket_slot_{x}_{y}_{state}"
RPA_BUCKET_AUTOMATION_ID_TEMPLATE = (
    "QApplication.MainWindow.QWidget.topCard.bucketColumnsHost."
    "bucketColumn__{x}.bucketPickRadio__{x}_{y}__{state}"
)
RPA_BUCKET_XPATH_TEMPLATE = (
    "/Group[@AutomationId='QApplication.MainWindow.QWidget' and @ClassName='QWidget']"
    "/Custom[@AutomationId='QApplication.MainWindow.QWidget.topCard' and @ClassName='QFrame']"
    "/Group[@AutomationId='QApplication.MainWindow.QWidget.topCard.bucketColumnsHost' and @ClassName='QWidget']"
    "/Group[@AutomationId='QApplication.MainWindow.QWidget.topCard.bucketColumnsHost.bucketColumn__{x}' and @ClassName='QWidget']"
    "/RadioButton[@AutomationId='QApplication.MainWindow.QWidget.topCard.bucketColumnsHost.bucketColumn__{x}.bucketPickRadio__{x}_{y}__{state}' and @Name='bucket_slot_{x}_{y}_{state}' and @ClassName='QRadioButton']"
)
RPA_BUCKET_EMPTY_RULE = "有桶且人数>0: Name/AutomationId 后缀为 enabled；0/容量或无桶: 后缀为 disabled"


@dataclass
class AppConfig:
    output_dir: str = DEFAULT_EXPORT_DIR
    source_dir: str = DEFAULT_SOURCE_DIR
    bucket_config_json: str = DEFAULT_BUCKET_CONFIG_JSON
    company_daily_limit: int = 10
    morning_prompt: str = DEFAULT_MORNING_PROMPT
    afternoon_prompt: str = DEFAULT_AFTERNOON_PROMPT
    extra_payload: Dict[str, Any] = field(default_factory=dict)


@dataclass
class AppState:
    config_path: str = str(DEFAULT_CONFIG_PATH)
    last_input_dir: str = ""


@dataclass
class ContactRecord:
    source_row: int
    company: str
    name: str
    title: str
    linkedin: str
    address: str
    emails: List[str]
    email_str: str = ""
    source_rows: List[int] = field(default_factory=list)


@dataclass
class CompanyBatch:
    company_key: str
    company_name: str
    file_path: Path
    contacts: List[ContactRecord] = field(default_factory=list)
    raw_row_count: int = 0
    valid_contact_count: int = 0
    merged_count: int = 0
    removed_b_count: int = 0


@dataclass
class BucketDef:
    bucket_key: str
    bucket_label: str
    weekday_key: str
    weekday_label: str
    mode: str
    mode_label: str
    target_date_str: str
    capacity: int


@dataclass
class AssignedEntry:
    row_id: str
    company_key: str
    company: str
    source_file: str
    name: str
    title: str
    linkedin: str
    address: str
    email_str: str
    bucket_key: str
    weekday_label: str
    mode: str
    mode_label: str
    target_date_str: str
    marketing_account: str = ""
    marketing_account_index: int = -1
    marketing_account_day_seq: int = 0
    marketing_account_global_seq: int = 0
    source_rows: List[int] = field(default_factory=list)


@dataclass
class UnassignedItem:
    company_key: str
    company: str
    source_file: str
    name: str
    title: str
    linkedin: str
    address: str
    email_str: str
    reason: str
    source_rows: List[int] = field(default_factory=list)


@dataclass
class Template2Record:
    row_id: str
    date_str: str = ""
    time_str: str = ""
    no_valid_time: bool = False


@dataclass
class Template3ExportRow:
    company: str
    name: str
    date_str: str
    time_str: str
    emails: List[str]


@dataclass
class ExceptionItem:
    category: str
    row_id: str
    detail: str


@dataclass
class ParseSummary:
    company_count: int = 0
    raw_row_count: int = 0
    valid_contact_count: int = 0
    merged_count: int = 0
    removed_b_count: int = 0
    assigned_count: int = 0
    unassigned_count: int = 0


@dataclass
class SourceLayout:
    header_row_index: int
    data_start_index: int
    name_col: Optional[int]
    title_col: Optional[int]
    linkedin_col: Optional[int]
    company_col: Optional[int]
    address_col: Optional[int]
    email_pairs: List[Tuple[Optional[int], int]]


@dataclass
class IndexRecord:
    file_name: str
    source_path: str = ""
    file_size: int = 0
    file_mtime: str = ""
    export_status: str = ""
    final_status: str = ""
    last_processed_date: str = ""
    completed_time: str = ""
    last_bucket: str = ""
    processed_count: int = 0
    total_count: int = 0
    deferred_count: int = 0
    has_deferred: bool = False
    deferred_path: str = ""
    earliest_reprocess_time: str = ""
    is_due: bool = False
    export_path: str = ""
    remark: str = ""


@dataclass
class StartupCandidate:
    file_path: Path
    file_name: str
    label: str
    checked: bool = True
    is_deferred: bool = False


class Store:
    def __init__(self) -> None:
        APP_DIR.mkdir(parents=True, exist_ok=True)
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        self._migrate_legacy_files()

    def _migrate_legacy_files(self) -> None:
        legacy_config_dir = LEGACY_APP_DIR / "configs"
        legacy_state_path = LEGACY_APP_DIR / "state.json"
        legacy_default_config_path = legacy_config_dir / "default_config.json"
        try:
            if legacy_default_config_path.exists() and not DEFAULT_CONFIG_PATH.exists():
                DEFAULT_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(legacy_default_config_path, DEFAULT_CONFIG_PATH)
            if legacy_state_path.exists() and not STATE_PATH.exists():
                shutil.copy2(legacy_state_path, STATE_PATH)
            if legacy_config_dir.exists():
                CONFIG_DIR.mkdir(parents=True, exist_ok=True)
                for item in legacy_config_dir.glob("*.json"):
                    target = CONFIG_DIR / item.name
                    if not target.exists():
                        shutil.copy2(item, target)
        except Exception:
            pass

    def ensure_default_config(self) -> Path:
        if not DEFAULT_CONFIG_PATH.exists():
            self.write_config(DEFAULT_CONFIG_PATH, AppConfig())
        return DEFAULT_CONFIG_PATH

    def load_state(self) -> AppState:
        self.ensure_default_config()
        if not STATE_PATH.exists():
            return AppState(config_path=str(DEFAULT_CONFIG_PATH))
        try:
            payload = json.loads(STATE_PATH.read_text(encoding="utf-8"))
            state = AppState()
            if isinstance(payload, dict):
                state.config_path = str(payload.get("config_path") or DEFAULT_CONFIG_PATH)
                state.last_input_dir = str(payload.get("last_input_dir") or "")
            return state
        except Exception:
            return AppState(config_path=str(DEFAULT_CONFIG_PATH))

    def save_state(self, state: AppState) -> None:
        payload = {
            "config_path": state.config_path,
            "last_input_dir": state.last_input_dir,
        }
        STATE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def load_config(self, config_path: Path) -> AppConfig:
        self.ensure_default_config()
        path = config_path.expanduser()
        if not path.exists():
            path = DEFAULT_CONFIG_PATH
        payload = json.loads(path.read_text(encoding="utf-8"))
        data = AppConfig()
        recognized = {"output_dir", "source_dir", "bucket_config_json", "company_daily_limit", "morning_prompt", "afternoon_prompt"}
        if isinstance(payload, dict):
            for key, value in payload.items():
                if key in recognized:
                    setattr(data, key, value)
                else:
                    data.extra_payload[key] = value
        parse_bucket_config(data.bucket_config_json)
        data.company_daily_limit = max(1, int(data.company_daily_limit))
        data.output_dir = str(data.output_dir or DEFAULT_EXPORT_DIR)
        data.source_dir = str(data.source_dir or DEFAULT_SOURCE_DIR)
        data.morning_prompt = str(data.morning_prompt or DEFAULT_MORNING_PROMPT)
        data.afternoon_prompt = str(data.afternoon_prompt or DEFAULT_AFTERNOON_PROMPT)
        return data

    def write_config(self, config_path: Path, config: AppConfig) -> None:
        payload = dict(config.extra_payload or {})
        payload.update({
            "output_dir": config.output_dir,
            "source_dir": config.source_dir,
            "bucket_config_json": config.bucket_config_json,
            "company_daily_limit": config.company_daily_limit,
            "morning_prompt": config.morning_prompt,
            "afternoon_prompt": config.afternoon_prompt,
        })
        config_path.parent.mkdir(parents=True, exist_ok=True)
        config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def import_config_copy(self, src_path: Path) -> Path:
        name = src_path.stem or "imported_config"
        target = build_non_overwriting_path(CONFIG_DIR, name, ".json")
        shutil.copy2(src_path, target)
        return target


class ClickableFrame(QFrame):
    clicked = Signal()

    def mousePressEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


class DropFrame(QFrame):
    filesDropped = Signal(list)
    clicked = Signal()

    def __init__(self, title: str, hint: str, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setCursor(Qt.PointingHandCursor)
        self.setObjectName("dropFrame")
        layout = QHBoxLayout(self)
        layout.setContentsMargins(18, 14, 18, 14)
        layout.setSpacing(14)

        badge = QLabel("XLSX")
        badge.setObjectName("dropBadge")
        badge.setAlignment(Qt.AlignCenter)
        badge.setFixedWidth(62)

        text_wrap = QVBoxLayout()
        text_wrap.setContentsMargins(0, 0, 0, 0)
        text_wrap.setSpacing(2)
        self.label = QLabel(title)
        self.label.setObjectName("dropText")
        self.sub_label = QLabel(hint)
        self.sub_label.setObjectName("dropSubText")
        text_wrap.addWidget(self.label)
        text_wrap.addWidget(self.sub_label)

        arrow = QLabel("点击导入")
        arrow.setObjectName("dropCta")
        arrow.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        for widget in (badge, self.label, self.sub_label, arrow):
            widget.setCursor(Qt.PointingHandCursor)
        layout.addWidget(badge)
        layout.addLayout(text_wrap, 1)
        layout.addWidget(arrow)

    def dragEnterEvent(self, event) -> None:  # type: ignore[override]
        if event.mimeData().hasUrls() and any(self._is_excel(url.toLocalFile()) for url in event.mimeData().urls()):
            event.acceptProposedAction()
            self.setProperty("dragging", True)
            self.style().unpolish(self)
            self.style().polish(self)
        else:
            event.ignore()

    def dragLeaveEvent(self, event) -> None:  # type: ignore[override]
        self.setProperty("dragging", False)
        self.style().unpolish(self)
        self.style().polish(self)
        super().dragLeaveEvent(event)

    def dropEvent(self, event) -> None:  # type: ignore[override]
        self.setProperty("dragging", False)
        self.style().unpolish(self)
        self.style().polish(self)
        paths: List[str] = []
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if self._is_excel(path):
                paths.append(path)
        if paths:
            self.filesDropped.emit(paths)
            event.acceptProposedAction()
        else:
            event.ignore()

    def mousePressEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)

    @staticmethod
    def _is_excel(path: str) -> bool:
        return Path(path).suffix.lower() in {".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"}


class CollapsibleSection(QFrame):
    def __init__(self, title: str, section_id: str, parent: Optional[QWidget] = None, default_open: bool = False) -> None:
        super().__init__(parent)
        self.section_id = section_id
        self.setObjectName(f"sectionBox__{section_id}")
        self.setProperty("sectionRole", "true")
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self.head = ClickableFrame()
        self.head.setObjectName(f"sectionHead__{section_id}")
        self.head.setProperty("sectionHeadRole", "true")
        self.head.setCursor(Qt.PointingHandCursor)
        self.head.setFixedHeight(26)
        self.head.clicked.connect(self.toggle)

        head_layout = QHBoxLayout(self.head)
        head_layout.setContentsMargins(6, 0, 6, 0)
        head_layout.setSpacing(4)

        self.toggle_btn = QToolButton()
        self.toggle_btn.setObjectName(f"sectionToggle__{section_id}")
        self.toggle_btn.setProperty("sectionToggleRole", "true")
        self.toggle_btn.setCursor(Qt.PointingHandCursor)
        self.toggle_btn.setCheckable(True)
        self.toggle_btn.setChecked(default_open)
        self.toggle_btn.clicked.connect(self._sync_open_state)
        self.toggle_btn.setToolButtonStyle(Qt.ToolButtonIconOnly)

        self.title_label = QLabel(title)
        self.title_label.setObjectName(f"sectionTitle__{section_id}")
        self.title_label.setProperty("sectionTitleRole", "true")
        self.title_label.setCursor(Qt.PointingHandCursor)

        self.meta_label = QLabel("")
        self.meta_label.setObjectName(f"sectionMeta__{section_id}")
        self.meta_label.setProperty("sectionMetaRole", "true")
        self.meta_label.setCursor(Qt.PointingHandCursor)

        self.fold_text_btn = QPushButton("收起" if default_open else "展开")
        self.fold_text_btn.setObjectName(f"sectionFold__{section_id}")
        self.fold_text_btn.setProperty("sectionFoldRole", "true")
        self.fold_text_btn.setCursor(Qt.PointingHandCursor)
        self.fold_text_btn.setFixedHeight(20)
        self.fold_text_btn.setFixedWidth(52)
        self.fold_text_btn.clicked.connect(self.toggle)

        head_layout.addWidget(self.toggle_btn)
        head_layout.addWidget(self.title_label)
        head_layout.addStretch(1)
        head_layout.addWidget(self.meta_label)
        head_layout.addWidget(self.fold_text_btn)
        root.addWidget(self.head)

        self.body = QWidget()
        self.body.setObjectName(f"sectionBody__{section_id}")
        self.body.setProperty("sectionBodyRole", "true")
        self.body_layout = QVBoxLayout(self.body)
        self.body_layout.setContentsMargins(6, 6, 6, 6)
        self.body_layout.setSpacing(4)
        self.body.setVisible(default_open)
        root.addWidget(self.body)

        self._sync_open_state()

    def toggle(self) -> None:
        self.toggle_btn.setChecked(not self.toggle_btn.isChecked())
        self._sync_open_state()

    def _sync_open_state(self) -> None:
        opened = self.toggle_btn.isChecked()
        self.toggle_btn.setArrowType(Qt.DownArrow if opened else Qt.RightArrow)
        self.body.setVisible(opened)
        self.fold_text_btn.setText("收起" if opened else "展开")

    def set_meta(self, text: str) -> None:
        self.meta_label.setText(text)
        self.meta_label.setVisible(bool(text))

    def set_open(self, opened: bool) -> None:
        self.toggle_btn.setChecked(opened)
        self._sync_open_state()


class SettingsDialog(QDialog):
    def __init__(self, current_config_path: Path, current_config: AppConfig, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("设置")
        self.resize(980, 760)
        self.store = Store()
        self.selected_config_path = current_config_path
        self.original_config = AppConfig(
            output_dir=current_config.output_dir,
            source_dir=current_config.source_dir,
            bucket_config_json=current_config.bucket_config_json,
            company_daily_limit=current_config.company_daily_limit,
            morning_prompt=current_config.morning_prompt,
            afternoon_prompt=current_config.afternoon_prompt,
            extra_payload=dict(current_config.extra_payload or {}),
        )
        self.edited_config = AppConfig(
            output_dir=current_config.output_dir,
            source_dir=current_config.source_dir,
            bucket_config_json=current_config.bucket_config_json,
            company_daily_limit=current_config.company_daily_limit,
            morning_prompt=current_config.morning_prompt,
            afternoon_prompt=current_config.afternoon_prompt,
            extra_payload=dict(current_config.extra_payload or {}),
        )

        root = QVBoxLayout(self)
        tabs = QTabWidget()
        tabs.setObjectName("settingsTabs")
        root.addWidget(tabs)

        settings_tab = QWidget()
        settings_layout = QVBoxLayout(settings_tab)
        settings_layout.setContentsMargins(12, 12, 12, 12)
        settings_layout.setSpacing(10)

        path_card = QFrame()
        path_card.setObjectName("sectionCard")
        path_card_layout = QVBoxLayout(path_card)
        path_card_layout.setContentsMargins(10, 10, 10, 10)
        path_card_layout.setSpacing(8)
        path_card_layout.addWidget(QLabel("当前配置文件路径"))
        self.current_path_edit = QPlainTextEdit(str(self.selected_config_path))
        self.current_path_edit.setReadOnly(True)
        self.current_path_edit.setFixedHeight(64)
        self.current_path_edit.setObjectName("previewEdit")
        self.current_path_edit.setProperty("previewRole", "true")
        path_card_layout.addWidget(self.current_path_edit)
        btn_row = QHBoxLayout()
        self.choose_btn = QPushButton("选择配置文件")
        self.import_btn = QPushButton("导入配置文件")
        self.choose_btn.clicked.connect(self.choose_config_file)
        self.import_btn.clicked.connect(self.import_config_file)
        btn_row.addWidget(self.choose_btn)
        btn_row.addWidget(self.import_btn)
        btn_row.addStretch(1)
        path_card_layout.addLayout(btn_row)
        settings_layout.addWidget(path_card)

        basic_card = QFrame()
        basic_card.setObjectName("sectionCard")
        basic_layout = QGridLayout(basic_card)
        basic_layout.setContentsMargins(10, 10, 10, 10)
        basic_layout.setHorizontalSpacing(10)
        basic_layout.setVerticalSpacing(8)
        basic_layout.addWidget(QLabel("导出目录"), 0, 0)
        self.output_dir_edit = QLineEdit(self.edited_config.output_dir)
        basic_layout.addWidget(self.output_dir_edit, 0, 1)
        self.output_dir_btn = QPushButton("浏览")
        self.output_dir_btn.clicked.connect(self.choose_output_dir)
        basic_layout.addWidget(self.output_dir_btn, 0, 2)
        basic_layout.addWidget(QLabel("数据源目录"), 1, 0)
        self.source_dir_edit = QLineEdit(self.edited_config.source_dir)
        basic_layout.addWidget(self.source_dir_edit, 1, 1)
        self.source_dir_btn = QPushButton("浏览")
        self.source_dir_btn.clicked.connect(self.choose_source_dir)
        basic_layout.addWidget(self.source_dir_btn, 1, 2)
        basic_layout.addWidget(QLabel("单公司单日上限"), 2, 0)
        self.company_daily_limit_edit = QLineEdit(str(self.edited_config.company_daily_limit))
        self.company_daily_limit_edit.setPlaceholderText("整数")
        basic_layout.addWidget(self.company_daily_limit_edit, 2, 1)
        settings_layout.addWidget(basic_card)

        bucket_card = QFrame()
        bucket_card.setObjectName("sectionCard")
        bucket_layout = QVBoxLayout(bucket_card)
        bucket_layout.setContentsMargins(10, 10, 10, 10)
        bucket_layout.setSpacing(8)
        bucket_layout.addWidget(QLabel("周桶配置（按天每账号容量）"))
        self.bucket_grid = QGridLayout()
        self.bucket_grid.setHorizontalSpacing(10)
        self.bucket_grid.setVerticalSpacing(8)
        self.bucket_grid.addWidget(QLabel("星期"), 0, 0)
        self.bucket_grid.addWidget(QLabel("日容量"), 0, 1)
        self.bucket_edits: Dict[str, QLineEdit] = {}
        daily_bucket_map = get_daily_bucket_capacity_map(self.edited_config.bucket_config_json)
        for row, day in enumerate(WEEKDAY_ORDER, start=1):
            self.bucket_grid.addWidget(QLabel(WEEKDAY_LABELS[day]), row, 0)
            edit = QLineEdit(str(daily_bucket_map.get(day, 0)))
            edit.setPlaceholderText("0")
            self.bucket_grid.addWidget(edit, row, 1)
            self.bucket_edits[day] = edit
        bucket_layout.addLayout(self.bucket_grid)
        settings_layout.addWidget(bucket_card)

        prompt_card = QFrame()
        prompt_card.setObjectName("sectionCard")
        prompt_layout = QGridLayout(prompt_card)
        prompt_layout.setContentsMargins(10, 10, 10, 10)
        prompt_layout.setHorizontalSpacing(10)
        prompt_layout.setVerticalSpacing(8)
        prompt_layout.addWidget(QLabel("上午提示词"), 0, 0)
        prompt_layout.addWidget(QLabel("下午提示词"), 0, 1)
        self.morning_prompt_edit = QPlainTextEdit(self.edited_config.morning_prompt)
        self.afternoon_prompt_edit = QPlainTextEdit(self.edited_config.afternoon_prompt)
        self.morning_prompt_edit.setObjectName("previewEdit")
        self.morning_prompt_edit.setProperty("previewRole", "true")
        self.afternoon_prompt_edit.setObjectName("previewEdit")
        self.afternoon_prompt_edit.setProperty("previewRole", "true")
        prompt_layout.addWidget(self.morning_prompt_edit, 1, 0)
        prompt_layout.addWidget(self.afternoon_prompt_edit, 1, 1)
        settings_layout.addWidget(prompt_card, 1)

        tabs.addTab(settings_tab, "设置")

        info_tab = QWidget()
        info_layout = QVBoxLayout(info_tab)
        info_layout.setContentsMargins(12, 12, 12, 12)
        self.info_edit = QPlainTextEdit(PROGRAM_INFO_TEXT)
        self.info_edit.setReadOnly(True)
        self.info_edit.setObjectName("previewEdit")
        self.info_edit.setProperty("previewRole", "true")
        info_layout.addWidget(self.info_edit)
        tabs.addTab(info_tab, "程序信息")

        foot = QHBoxLayout()
        self.save_tip_label = QLabel("保存后将直接写入当前配置 JSON")
        self.save_tip_label.setObjectName("metaLine")
        foot.addWidget(self.save_tip_label)
        foot.addStretch(1)
        ok_btn = QPushButton("保存并关闭")
        cancel_btn = QPushButton("取消")
        ok_btn.clicked.connect(self.save_and_accept)
        cancel_btn.clicked.connect(self.reject)
        foot.addWidget(ok_btn)
        foot.addWidget(cancel_btn)
        root.addLayout(foot)

    def choose_output_dir(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "选择导出目录", self.output_dir_edit.text().strip() or str(Path.home()))
        if folder:
            self.output_dir_edit.setText(folder)

    def choose_source_dir(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "选择数据源目录", self.source_dir_edit.text().strip() or str(Path.home()))
        if folder:
            self.source_dir_edit.setText(folder)

    def choose_config_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "选择配置文件", str(self.selected_config_path.parent), "JSON Files (*.json)")
        if not path:
            return
        try:
            config = self.store.load_config(Path(path))
        except Exception as exc:
            QMessageBox.warning(self, "配置文件无效", str(exc))
            return
        self.selected_config_path = Path(path)
        self.current_path_edit.setPlainText(str(self.selected_config_path))
        self._load_form_from_config(config)

    def import_config_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "导入配置文件", str(Path.home()), "JSON Files (*.json)")
        if not path:
            return
        try:
            copied = self.store.import_config_copy(Path(path))
            config = self.store.load_config(copied)
        except Exception as exc:
            QMessageBox.warning(self, "导入失败", str(exc))
            return
        self.selected_config_path = copied
        self.current_path_edit.setPlainText(str(self.selected_config_path))
        self._load_form_from_config(config)
        QMessageBox.information(self, "导入完成", f"已导入：\n{copied}")

    def _load_form_from_config(self, config: AppConfig) -> None:
        self.edited_config = AppConfig(
            output_dir=config.output_dir,
            source_dir=config.source_dir,
            bucket_config_json=config.bucket_config_json,
            company_daily_limit=config.company_daily_limit,
            morning_prompt=config.morning_prompt,
            afternoon_prompt=config.afternoon_prompt,
            extra_payload=dict(config.extra_payload or {}),
        )
        self.output_dir_edit.setText(self.edited_config.output_dir)
        self.source_dir_edit.setText(self.edited_config.source_dir)
        self.company_daily_limit_edit.setText(str(self.edited_config.company_daily_limit))
        daily_bucket_map = get_daily_bucket_capacity_map(self.edited_config.bucket_config_json)
        for day in WEEKDAY_ORDER:
            self.bucket_edits[day].setText(str(daily_bucket_map.get(day, 0)))
        self.morning_prompt_edit.setPlainText(self.edited_config.morning_prompt)
        self.afternoon_prompt_edit.setPlainText(self.edited_config.afternoon_prompt)

    def _collect_config_from_form(self) -> AppConfig:
        output_dir = self.output_dir_edit.text().strip() or DEFAULT_EXPORT_DIR
        source_dir = self.source_dir_edit.text().strip()
        try:
            company_daily_limit = max(1, int((self.company_daily_limit_edit.text() or "").strip()))
        except Exception as exc:
            raise ValueError("单公司单日上限必须是整数。") from exc

        bucket_payload: Dict[str, int] = {}
        for day in WEEKDAY_ORDER:
            raw = self.bucket_edits[day].text().strip()
            try:
                value = int(raw or "0")
            except Exception as exc:
                raise ValueError(f"{WEEKDAY_LABELS[day]} 日容量必须是整数。") from exc
            if value < 0:
                raise ValueError(f"{WEEKDAY_LABELS[day]} 日容量不能小于 0。")
            bucket_payload[day] = value

        bucket_config_json = json.dumps(bucket_payload, ensure_ascii=False, indent=2)
        parse_bucket_config(bucket_config_json)

        morning_prompt = self.morning_prompt_edit.toPlainText().strip()
        afternoon_prompt = self.afternoon_prompt_edit.toPlainText().strip()
        if not morning_prompt:
            raise ValueError("上午提示词不能为空。")
        if not afternoon_prompt:
            raise ValueError("下午提示词不能为空。")

        return AppConfig(
            output_dir=output_dir,
            source_dir=source_dir,
            bucket_config_json=bucket_config_json,
            company_daily_limit=company_daily_limit,
            morning_prompt=morning_prompt,
            afternoon_prompt=afternoon_prompt,
            extra_payload=dict(self.edited_config.extra_payload or {}),
        )

    def save_and_accept(self) -> None:
        try:
            self.edited_config = self._collect_config_from_form()
            self.store.write_config(self.selected_config_path, self.edited_config)
        except Exception as exc:
            QMessageBox.warning(self, "保存失败", str(exc))
            return
        self.accept()



class DeferredSelectionDialog(QDialog):
    def __init__(self, items: List[StartupCandidate], parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("顺延待处理清单")
        self.resize(760, 520)
        self.items = items
        root = QVBoxLayout(self)
        hint = QLabel("以下顺延文件已到可处理时间。勾选后将随启动自动载入。")
        hint.setObjectName("metaLine")
        root.addWidget(hint)

        action_row = QHBoxLayout()
        select_all_btn = QPushButton("全选")
        clear_all_btn = QPushButton("全不选")
        select_all_btn.clicked.connect(self.select_all)
        clear_all_btn.clicked.connect(self.clear_all)
        action_row.addWidget(select_all_btn)
        action_row.addWidget(clear_all_btn)
        action_row.addStretch(1)
        root.addLayout(action_row)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setObjectName("deferredScroll")
        holder = QWidget()
        self.scroll_layout = QVBoxLayout(holder)
        self.scroll_layout.setContentsMargins(10, 10, 10, 10)
        self.scroll_layout.setSpacing(8)
        self.checkboxes: List[Tuple[StartupCandidate, QCheckBox]] = []
        for item in items:
            text = f"{item.file_name} | {item.label}"
            checkbox = QCheckBox(text)
            checkbox.setChecked(item.checked)
            self.scroll_layout.addWidget(checkbox)
            self.checkboxes.append((item, checkbox))
        self.scroll_layout.addStretch(1)
        self.scroll.setWidget(holder)
        root.addWidget(self.scroll, 1)

        foot = QHBoxLayout()
        foot.addStretch(1)
        ok_btn = QPushButton("确定")
        cancel_btn = QPushButton("取消")
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        foot.addWidget(ok_btn)
        foot.addWidget(cancel_btn)
        root.addLayout(foot)

    def select_all(self) -> None:
        for _, checkbox in self.checkboxes:
            checkbox.setChecked(True)

    def clear_all(self) -> None:
        for _, checkbox in self.checkboxes:
            checkbox.setChecked(False)

    def selected_paths(self) -> List[Path]:
        result: List[Path] = []
        for item, checkbox in self.checkboxes:
            if checkbox.isChecked():
                result.append(item.file_path)
        return result


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle(APP_NAME)
        self.resize(1320, 900)
        self.setAcceptDrops(True)

        self.store = Store()
        self.state = self.store.load_state()
        self.current_config_path = Path(self.state.config_path).expanduser()
        try:
            self.config = self.store.load_config(self.current_config_path)
        except Exception:
            self.current_config_path = self.store.ensure_default_config()
            self.config = self.store.load_config(self.current_config_path)
            self.state.config_path = str(self.current_config_path)
            self.store.save_state(self.state)

        self.current_source_paths: List[Path] = []
        self.current_export_path: Optional[Path] = None
        self.current_deferred_export_paths: List[Path] = []
        self.current_bucket_key: str = ""
        self.current_batch_code: str = ""
        self.current_batch_label: str = ""
        self.template3_result_map: Dict[str, Tuple[str, str, str]] = {}
        self.bucket_button_map: Dict[str, QRadioButton] = {}
        self.allowed_end_cn_date_str: str = ""
        self.current_state_text: str = "未载入文件"
        self.summary = ParseSummary()
        self.company_batches: List[CompanyBatch] = []
        self.bucket_defs: List[BucketDef] = []
        self.assigned_entries: List[AssignedEntry] = []
        self.unassigned_items: List[UnassignedItem] = []
        self.template2_records: Dict[str, Template2Record] = {}
        self.template2_generated_at_utc: Optional[datetime] = None
        self.template2_source_bucket_key: str = ""
        self.template3_lines: List[str] = []
        self.template3_exports: List[Template3ExportRow] = []
        self.batch_template2_records: Dict[str, Template2Record] = {}
        self.batch_template3_exports: Dict[str, Template3ExportRow] = {}
        self.batch_exceptions: List[ExceptionItem] = []
        self.exceptions: List[ExceptionItem] = []
        self.index_records: Dict[str, IndexRecord] = {}
        self.startup_boot_done = False

        self._build_ui()
        self._apply_styles()
        self._rebuild_empty_buckets_from_config()
        self._update_file_info_labels()
        self._update_ui_state()
        QTimer.singleShot(0, self._startup_bootstrap)

    def _build_ui(self) -> None:
        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 10, 12, 12)
        root.setSpacing(8)
        self.setCentralWidget(central)

        top = QFrame()
        top.setObjectName("topCard")
        top_layout = QVBoxLayout(top)
        top_layout.setContentsMargins(14, 12, 14, 10)
        top_layout.setSpacing(6)

        title_row = QHBoxLayout()
        title = QLabel("联系人清洗 · 周分桶")
        title.setObjectName("titleLabel")
        self.settings_btn = QPushButton("设置")
        self.settings_btn.setObjectName("settingsSoloBtn")
        self.settings_btn.setFixedWidth(86)
        self.settings_btn.clicked.connect(self.open_settings)
        self.state_badge = QLabel("未载入文件")
        self.state_badge.setObjectName("stateBadge")
        title_row.addWidget(title)
        title_row.addStretch(1)
        title_row.addWidget(self.settings_btn)
        title_row.addWidget(self.state_badge)
        top_layout.addLayout(title_row)

        row1 = QHBoxLayout()
        self.file_row_label = QLabel("文件：未选择")
        self.file_row_label.setObjectName("metaLine")
        self.config_row_label = QLabel("配置：default_config.json")
        self.config_row_label.setObjectName("metaLine")
        row1.addWidget(self.file_row_label, 1)
        row1.addWidget(self.config_row_label, 1)
        top_layout.addLayout(row1)

        row2 = QHBoxLayout()
        self.bucket_info_label = QLabel("当前桶：未分配")
        self.bucket_info_label.setObjectName("metaLine")
        row2.addWidget(self.bucket_info_label, 1)
        self.processing_window_label = QLabel("处理空间：未分配")
        self.processing_window_label.setObjectName("metaLine")
        row2.addWidget(self.processing_window_label)
        self.template2_base_label = QLabel("模板2基准：未生成")
        self.template2_base_label.setObjectName("metaLine")
        row2.addWidget(self.template2_base_label)
        top_layout.addLayout(row2)

        self.bucket_block_title = QLabel("分桶")
        self.bucket_block_title.setObjectName("bucketBlockTitle")
        top_layout.addWidget(self.bucket_block_title)
        self.bucket_columns_host = QWidget()
        self.bucket_columns_host.setObjectName("bucketColumnsHost")
        self.bucket_columns_layout = QHBoxLayout(self.bucket_columns_host)
        self.bucket_columns_layout.setContentsMargins(0, 0, 0, 0)
        self.bucket_columns_layout.setSpacing(8)
        top_layout.addWidget(self.bucket_columns_host)
        self.bucket_button_group = QButtonGroup(self)
        self.bucket_button_group.setExclusive(True)
        root.addWidget(top)

        strip = QFrame()
        strip.setObjectName("summaryCard")
        strip_layout = QHBoxLayout(strip)
        strip_layout.setContentsMargins(12, 3, 12, 3)
        self.summary_label = QLabel("公司0|数据0|有效0|去重0|已分配0|未分配0|异常0")
        self.summary_label.setObjectName("summaryLine")
        strip_layout.addWidget(self.summary_label, 1)
        root.addWidget(strip)

        self.drop_frame = DropFrame("导入公司文件", "点击或拖拽一个或多个 Excel 到此处")
        self.drop_frame.clicked.connect(self.choose_multiple_files)
        self.drop_frame.filesDropped.connect(self.handle_dropped_files)
        self.drop_frame.setFixedHeight(118)
        root.addWidget(self.drop_frame)

        ops1 = QFrame()
        ops1.setObjectName("buttonBarImport")
        ops1.setProperty("buttonBarRole", "true")
        ops1_layout = QGridLayout(ops1)
        ops1_layout.setContentsMargins(8, 8, 8, 8)
        self.single_btn = QPushButton("导入单个公司")
        self.single_btn.setObjectName("buttonImportSingle")
        self.single_btn.clicked.connect(self.choose_single_file)
        self.multi_btn = QPushButton("导入多个公司")
        self.multi_btn.setObjectName("buttonImportMultiple")
        self.multi_btn.clicked.connect(self.choose_multiple_files)
        self.rebuild_btn = QPushButton("重新分桶")
        self.rebuild_btn.setObjectName("buttonRebuildBucket")
        self.rebuild_btn.clicked.connect(self.rebuild_from_current_files)
        self.copy_template1_btn = QPushButton("复制模板1")
        self.copy_template1_btn.setObjectName("buttonCopyTemplate1")
        self.copy_template1_btn.clicked.connect(self.copy_template1)
        for idx, btn in enumerate([self.single_btn, self.multi_btn, self.rebuild_btn, self.copy_template1_btn]):
            btn.setProperty("baseText", btn.text())
            btn.setMinimumHeight(32)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            ops1_layout.addWidget(btn, 0, idx)
        root.addWidget(ops1)

        ops2 = QFrame()
        ops2.setObjectName("buttonBarTemplate")
        ops2.setProperty("buttonBarRole", "true")
        ops2_layout = QGridLayout(ops2)
        ops2_layout.setContentsMargins(8, 8, 8, 8)
        self.copy_template2_input_btn = QPushButton("复制模板2输入稿")
        self.copy_template2_input_btn.setObjectName("buttonCopyTemplate2Input")
        self.copy_template2_input_btn.clicked.connect(self.copy_current_bucket_prompt)
        self.parse_template2_btn = QPushButton("解析模板2")
        self.parse_template2_btn.setObjectName("buttonParseTemplate2")
        self.parse_template2_btn.clicked.connect(self.generate_template3)
        self.clear_template2_btn = QPushButton("清空模板2")
        self.clear_template2_btn.setObjectName("buttonClearTemplate2")
        self.clear_template2_btn.clicked.connect(self.clear_template2)
        self.copy_template3_btn = QPushButton("复制模板3")
        self.copy_template3_btn.setObjectName("buttonCopyTemplate3")
        self.copy_template3_btn.clicked.connect(self.copy_template3)
        self.export_btn = QPushButton("导出结果")
        self.export_btn.setObjectName("buttonExportResult")
        self.export_btn.clicked.connect(self.export_current_workbook)
        for idx, btn in enumerate([
            self.copy_template2_input_btn,
            self.parse_template2_btn,
            self.clear_template2_btn,
            self.copy_template3_btn,
            self.export_btn,
        ]):
            btn.setProperty("baseText", btn.text())
            btn.setMinimumHeight(32)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            ops2_layout.addWidget(btn, 0, idx)
        root.addWidget(ops2)

        ops3 = QFrame()
        ops3.setObjectName("buttonBarLogs")
        ops3.setProperty("buttonBarRole", "true")
        ops3_layout = QGridLayout(ops3)
        ops3_layout.setContentsMargins(8, 8, 8, 8)
        self.copy_company_log_btn = QPushButton("复制公司日志")
        self.copy_company_log_btn.setObjectName("buttonCopyCompanyLog")
        self.copy_company_log_btn.clicked.connect(self.copy_company_log_rows)
        self.copy_contact_log_btn = QPushButton("复制联系人日志")
        self.copy_contact_log_btn.setObjectName("buttonCopyContactLog")
        self.copy_contact_log_btn.clicked.connect(self.copy_contact_log_rows)
        for idx, btn in enumerate([self.copy_company_log_btn, self.copy_contact_log_btn]):
            btn.setProperty("baseText", btn.text())
            btn.setMinimumHeight(32)
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            ops3_layout.addWidget(btn, 0, idx)
        root.addWidget(ops3)

        preview_wrap = QFrame()
        preview_wrap.setObjectName("previewWrap")
        preview_wrap.setProperty("previewWrapRole", "true")
        preview_layout = QVBoxLayout(preview_wrap)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        preview_layout.setSpacing(6)

        self.template1_section = CollapsibleSection("模板1（当前桶）", "template1", default_open=False)
        self.template1_preview = self._build_preview_editor(120, placeholder="载入文件并选择当前桶后显示", object_name="previewTemplate1")
        self.template1_section.body_layout.addWidget(self.template1_preview)
        preview_layout.addWidget(self.template1_section)

        self.template2_section = CollapsibleSection("模板2粘贴区", "template2", default_open=True)
        self.template2_input = self._build_input_editor(190, placeholder="row_id|YYYY/MM/DD|HH:MM 或 row_id|NO_VALID_TIME", object_name="inputTemplate2")
        self.template2_input.textChanged.connect(self._on_template2_text_changed)
        self.template2_section.body_layout.addWidget(self.template2_input)
        preview_layout.addWidget(self.template2_section)

        self.template3_section = CollapsibleSection("模板3（当前桶）", "template3", default_open=False)
        self.template3_preview = self._build_preview_editor(120, placeholder="解析后显示", object_name="previewTemplate3")
        self.template3_section.body_layout.addWidget(self.template3_preview)
        preview_layout.addWidget(self.template3_section)

        self.unassigned_section = CollapsibleSection("未分配清单", "unassigned", default_open=False)
        self.unassigned_preview = self._build_preview_editor(110, placeholder="暂无未分配", object_name="previewUnassigned")
        self.unassigned_section.body_layout.addWidget(self.unassigned_preview)
        preview_layout.addWidget(self.unassigned_section)

        self.exception_section = CollapsibleSection("异常清单", "exception", default_open=False)
        self.exception_preview = self._build_preview_editor(110, placeholder="暂无异常", object_name="previewException")
        self.exception_section.body_layout.addWidget(self.exception_preview)
        preview_layout.addWidget(self.exception_section)

        preview_layout.addStretch(1)
        root.addWidget(preview_wrap, 1)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("未载入文件")

    def _build_preview_editor(self, height: int, placeholder: str = "", object_name: str = "previewEdit") -> QPlainTextEdit:
        editor = QPlainTextEdit()
        editor.setReadOnly(True)
        editor.setWordWrapMode(QTextOption.NoWrap)
        editor.setPlaceholderText(placeholder)
        editor.setFixedHeight(height)
        editor.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        editor.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        editor.setObjectName(object_name)
        editor.setProperty("previewRole", "true")
        return editor

    def _build_input_editor(self, height: int, placeholder: str = "", object_name: str = "inputEdit") -> QPlainTextEdit:
        editor = QPlainTextEdit()
        editor.setWordWrapMode(QTextOption.NoWrap)
        editor.setPlaceholderText(placeholder)
        editor.setFixedHeight(height)
        editor.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        editor.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        editor.setObjectName(object_name)
        editor.setProperty("inputRole", "true")
        return editor

    def _apply_styles(self) -> None:
        self.setStyleSheet(
            """
            QMainWindow, QWidget {
                background: #0A1118;
                color: #EAF0F6;
                font-family: "Microsoft YaHei", "PingFang SC", "Segoe UI";
                font-size: 13px;
            }
            QFrame#topCard, QFrame[buttonBarRole="true"], QFrame[sectionRole="true"], QFrame#summaryCard {
                background: #0F1720;
                border: 1px solid #233140;
                border-radius: 0;
            }
            QFrame#dropFrame {
                background: #0E1620;
                border: 1px dashed #35506B;
            }
            QFrame[sectionHeadRole="true"] {
                background: #0F1720;
                border-bottom: 1px solid #233140;
            }
            QFrame[sectionHeadRole="true"]:hover, QFrame#dropFrame:hover, QFrame#dropFrame[dragging="true"] {
                background: #132030;
            }
            QLabel#titleLabel { font-size: 20px; font-weight: 900; }
            QLabel#metaLine, QLabel[sectionMetaRole="true"], QLabel#dropSubText { color: #9DB0C3; font-weight: 700; }
            QLabel#summaryLine { color: #D7E1EC; font-size: 11px; font-weight: 700; }
            QLabel#dropBadge {
                color: #DCE8F4;
                border: 1px solid #35506B;
                font-weight: 900;
                padding: 16px 0;
            }
            QLabel#dropText, QLabel[sectionTitleRole="true"] { font-weight: 800; color: #E6EDF5; }
            QLabel#bucketBlockTitle {
                color: #D7E1EC;
                font-size: 11px;
                font-weight: 900;
                letter-spacing: 1px;
                padding-top: 2px;
            }
            QWidget#bucketColumnsHost {
                background: transparent;
            }
            QWidget[bucketColumnRole="true"] {
                background: transparent;
            }
            QFrame#bucketSeparator {
                background: #3D5A76;
                min-width: 1px;
                max-width: 1px;
            }
            QRadioButton[bucketRole="true"] {
                background: transparent;
                color: #EAF0F6;
                spacing: 7px;
                padding: 1px 6px 1px 4px;
                min-height: 18px;
                font-weight: 800;
                border: 0;
            }
            QRadioButton[bucketRole="true"]:hover {
                background: #132030;
            }
            QRadioButton[bucketRole="true"][hasData="false"] {
                color: #8EA3B6;
            }
            QRadioButton[bucketRole="true"]:disabled {
                color: #66798B;
                background: transparent;
            }
            QRadioButton[bucketRole="true"]:disabled::indicator {
                border: 1px solid #314255;
                background: #0B1219;
            }
            QRadioButton[bucketRole="true"][isFull="true"] {
                color: #FFD7AF;
            }
            QRadioButton[bucketRole="true"]::indicator {
                width: 11px;
                height: 11px;
                border: 1px solid #4A627A;
                border-radius: 0;
                background: #0A1118;
            }
            QRadioButton[bucketRole="true"]::indicator:checked {
                background: #7FB5EA;
                border: 1px solid #7FB5EA;
            }
            QLabel#dropText { font-size: 16px; }
            QLabel#dropCta { color: #A6C0DA; font-size: 11px; font-weight: 800; }
            QPushButton, QLineEdit, QPlainTextEdit {
                background: #0B1219;
                color: #EAF0F6;
                border: 1px solid #314255;
                border-radius: 0;
            }
            QPushButton {
                padding: 7px 10px;
                font-weight: 800;
            }
            QPushButton:hover { background: #132030; border: 1px solid #45627F; }
            QPushButton:disabled { color: #637587; border: 1px solid #233140; }
            QPushButton#settingsSoloBtn { background: transparent; }
            QLabel#stateBadge {
                background: #132030;
                border: 1px solid #35506B;
                padding: 2px 10px;
                font-weight: 800;
            }
            QPlainTextEdit[previewRole="true"], QPlainTextEdit[inputRole="true"], QPlainTextEdit#previewEdit, QPlainTextEdit#inputEdit, QLineEdit {
                padding: 6px;
            }
            QToolButton { background: transparent; border: 0; }
            QPushButton[sectionFoldRole="true"] {
                background: transparent;
                color: #B7CCE0;
                border: 1px solid #2E4254;
                padding: 0 6px;
                font-size: 11px;
                font-weight: 900;
            }
            QTabWidget#settingsTabs::pane {
                border: 1px solid #35506B;
                top: -1px;
                background: #0F1720;
            }
            QTabWidget#settingsTabs QTabBar::tab {
                min-width: 124px;
                min-height: 34px;
                padding: 6px 12px;
                margin-right: 2px;
                background: #101821;
                color: #8EA4BA;
                border: 1px solid #2A3B4C;
                font-weight: 900;
            }
            QTabWidget#settingsTabs QTabBar::tab:selected {
                background: #18314B;
                color: #F2F7FC;
                border: 1px solid #5E89B2;
            }
            QTabWidget#settingsTabs QTabBar::tab:hover:!selected {
                background: #152432;
                color: #CFE0F0;
            }
            QStatusBar { background: #0F1720; color: #8699AC; }
            """
        )

    def _shorten_middle(self, text: str, max_len: int = 42) -> str:
        if len(text) <= max_len:
            return text
        head = max_len // 2 - 1
        tail = max_len - head - 1
        return f"{text[:head]}…{text[-tail:]}"

    def _rebuild_empty_buckets_from_config(self) -> None:
        try:
            bucket_caps = expand_bucket_caps_for_accounts(parse_bucket_config(self.config.bucket_config_json))
            self.bucket_defs = build_window_bucket_defs(get_current_cn_datetime(), bucket_caps)
            self.allowed_end_cn_date_str = get_allowed_end_cn_date_str(self.bucket_defs)
        except Exception:
            self.bucket_defs = []
            self.allowed_end_cn_date_str = ""
        self.current_bucket_key = self._get_first_non_empty_bucket_key()

    def _get_bucket_count_map(self) -> Dict[str, int]:
        bucket_count_map = {item.bucket_key: 0 for item in self.bucket_defs}
        for entry in self.assigned_entries:
            if entry.bucket_key in bucket_count_map:
                bucket_count_map[entry.bucket_key] += 1
        return bucket_count_map

    def _get_first_non_empty_bucket_key(self, bucket_count_map: Optional[Dict[str, int]] = None) -> str:
        active_count_map = bucket_count_map if bucket_count_map is not None else self._get_bucket_count_map()
        for bucket in self.bucket_defs:
            if active_count_map.get(bucket.bucket_key, 0) > 0:
                return bucket.bucket_key
        return ""

    def _sync_current_bucket_key_with_data(self, bucket_count_map: Optional[Dict[str, int]] = None) -> None:
        active_count_map = bucket_count_map if bucket_count_map is not None else self._get_bucket_count_map()
        enabled_keys = {bucket.bucket_key for bucket in self.bucket_defs if active_count_map.get(bucket.bucket_key, 0) > 0}
        if not enabled_keys:
            self.current_bucket_key = ""
            return
        if self.current_bucket_key not in enabled_keys:
            self.current_bucket_key = self._get_first_non_empty_bucket_key(active_count_map)


    def _update_file_info_labels(self) -> None:
        file_text = f"{len(self.current_source_paths)} 个文件" if self.current_source_paths else "未选择"
        if len(self.current_source_paths) == 1:
            file_text = self.current_source_paths[0].name
        config_text = self.current_config_path.name if self.current_config_path else "未设置"
        self.file_row_label.setText(f"文件：{self._shorten_middle(file_text, 34)}")
        self.config_row_label.setText(f"配置：{self._shorten_middle(config_text, 34)}")

    def get_effective_source_dir(self) -> Optional[Path]:
        raw = str(self.config.source_dir or self.state.last_input_dir or "").strip()
        if not raw:
            return None
        path = Path(raw).expanduser()
        return path

    def get_index_dir(self) -> Path:
        return Path(self.config.output_dir).expanduser() / INDEX_SUBDIR_NAME

    def _load_index_records(self) -> Dict[str, IndexRecord]:
        index_dir = self.get_index_dir()
        index_dir.mkdir(parents=True, exist_ok=True)
        json_path = index_dir / INDEX_JSON_NAME
        xlsx_path = index_dir / INDEX_XLSX_NAME
        records = sync_and_load_index_records(json_path, xlsx_path)
        records = refresh_index_due_flags(records, get_current_cn_datetime())
        write_index_json(json_path, records)
        write_index_xlsx(xlsx_path, records)
        return records

    def _save_index_records(self) -> None:
        index_dir = self.get_index_dir()
        index_dir.mkdir(parents=True, exist_ok=True)
        json_path = index_dir / INDEX_JSON_NAME
        xlsx_path = index_dir / INDEX_XLSX_NAME
        self.index_records = refresh_index_due_flags(self.index_records, get_current_cn_datetime())
        write_index_json(json_path, self.index_records)
        write_index_xlsx(xlsx_path, self.index_records)

    def _startup_bootstrap(self) -> None:
        if self.startup_boot_done:
            return
        self.startup_boot_done = True
        try:
            self.index_records = self._load_index_records()
            self.status_bar.showMessage("启动完成：未自动载入旧批次，桶已清空")
        except Exception as exc:
            self.status_bar.showMessage(f"启动扫描失败：{exc}")

    def _startup_auto_load_source_files(self) -> None:
        source_dir = self.get_effective_source_dir()
        if source_dir is None or not source_dir.exists():
            return
        startup_candidates, skipped_messages = collect_startup_candidates(source_dir, self.index_records)
        due_items = [item for item in startup_candidates if item.is_deferred]
        selected_due_paths: List[Path] = []
        if due_items:
            dialog = DeferredSelectionDialog(due_items, self)
            if dialog.exec() == QDialog.Accepted:
                selected_due_paths = dialog.selected_paths()
            else:
                selected_due_paths = []
        selected_paths = [item.file_path for item in startup_candidates if not item.is_deferred]
        selected_paths.extend(selected_due_paths)
        deduped_paths: List[Path] = []
        seen = set()
        for path in selected_paths:
            if path.exists() and str(path.resolve()) not in seen:
                seen.add(str(path.resolve()))
                deduped_paths.append(path)
        if deduped_paths:
            self.load_source_files(deduped_paths)
            loaded_names = "，".join(path.name for path in deduped_paths[:4])
            if len(deduped_paths) > 4:
                loaded_names += f" 等{len(deduped_paths)}个"
            extra = f"；已跳过 {len(skipped_messages)} 个已处理文件" if skipped_messages else ""
            self.status_bar.showMessage(f"启动自动载入：{loaded_names}{extra}")
        elif skipped_messages:
            self.status_bar.showMessage(f"启动扫描完成：已跳过 {len(skipped_messages)} 个已处理文件")

    def _set_state_badge(self) -> None:
        self.state_badge.setText(self.current_state_text)
        if self.current_state_text == "异常":
            style = "background: rgba(244, 99, 99, 0.16); border: 1px solid rgba(244, 99, 99, 0.4); color: #FFD4D4;"
        elif self.current_state_text in {"模板3已生成", "已导出"}:
            style = "background: rgba(92, 184, 122, 0.16); border: 1px solid rgba(92, 184, 122, 0.34); color: #D7F5DE;"
        elif self.current_state_text in {"已就绪", "模板2已粘贴", "模板2输入稿已生成"}:
            style = "background: rgba(119,183,255,0.14); border: 1px solid #35527A; color: #DDEBFF;"
        else:
            style = "background: rgba(255, 186, 73, 0.14); border: 1px solid rgba(255, 186, 73, 0.34); color: #FFE2A8;"
        self.state_badge.setStyleSheet(style)

    def _clear_layout(self, layout) -> None:
        while layout.count():
            item = layout.takeAt(0)
            child_layout = item.layout()
            child_widget = item.widget()
            if child_layout is not None:
                self._clear_layout(child_layout)
            if child_widget is not None:
                child_widget.deleteLater()

    def _build_bucket_buttons(self) -> None:
        self.bucket_button_map = {}
        for button in self.bucket_button_group.buttons():
            self.bucket_button_group.removeButton(button)
        self._clear_layout(self.bucket_columns_layout)

        bucket_count_map = self._get_bucket_count_map()
        self._sync_current_bucket_key_with_data(bucket_count_map)

        for col in range(FIXED_BUCKET_COLUMNS):
            column_no = col + 1
            column_widget = QWidget()
            column_widget.setObjectName(f"bucketColumn__{column_no}")
            column_widget.setProperty("bucketColumnRole", "true")
            column_layout = QVBoxLayout(column_widget)
            column_layout.setContentsMargins(0, 0, 0, 0)
            column_layout.setSpacing(1)

            for row in range(FIXED_BUCKET_ROWS):
                row_no = row + 1
                absolute_index = bucket_slot_absolute_index(column_no, row_no)
                slot_key = f"{column_no}_{row_no}"
                radio = QRadioButton()
                radio.setProperty("bucketRole", "true")
                radio.setProperty("stableBucketSlot", slot_key)
                radio.setProperty("stableBucketColumn", column_no)
                radio.setProperty("stableBucketRow", row_no)
                radio.clicked.connect(self._on_bucket_button_clicked)

                bucket = self.bucket_defs[absolute_index - 1] if absolute_index <= len(self.bucket_defs) else None
                count = bucket_count_map.get(bucket.bucket_key, 0) if bucket is not None else 0
                slot_has_clickable_data = bucket is not None and count > 0
                state_suffix = "enabled" if slot_has_clickable_data else "disabled"
                radio.setObjectName(build_rpa_bucket_automation_id(column_no, row_no, state_suffix).split(".")[-1])
                radio.setProperty("rpaBucketXPath", build_rpa_bucket_xpath(column_no, row_no, state_suffix))
                radio.setAccessibleName(build_rpa_bucket_name(column_no, row_no, state_suffix))
                if bucket is not None:
                    radio.setText(f"{bucket.target_date_str} {bucket.bucket_label} | {count}/{bucket.capacity}")
                    radio.setProperty("bucketKey", bucket.bucket_key)
                    radio.setProperty("hasData", "true" if count > 0 else "false")
                    radio.setProperty("isFull", "true" if count >= bucket.capacity else "false")
                    radio.setProperty("rpaState", state_suffix)
                    radio.setProperty("rpaDisabled", "false" if slot_has_clickable_data else "true")
                    radio.setAccessibleDescription(
                        f"bucket|slot={column_no},{row_no}|state={state_suffix}|count={count}|{bucket.target_date_str}|{bucket.bucket_label}"
                    )
                    radio.setToolTip(bucket.bucket_key if slot_has_clickable_data else f"{bucket.bucket_key} | 当前无数据")
                    radio.setChecked(slot_has_clickable_data and bucket.bucket_key == self.current_bucket_key)
                    radio.setEnabled(slot_has_clickable_data)
                    self.bucket_button_map[bucket.bucket_key] = radio
                    self.bucket_button_group.addButton(radio)
                else:
                    radio.setText(f"空槽 {column_no},{row_no} | 无桶")
                    radio.setProperty("bucketKey", "")
                    radio.setProperty("hasData", "false")
                    radio.setProperty("isFull", "false")
                    radio.setProperty("rpaState", "disabled")
                    radio.setProperty("rpaDisabled", "true")
                    radio.setAccessibleDescription(f"bucket|slot={column_no},{row_no}|state=disabled|empty")
                    radio.setToolTip("无桶")
                    radio.setChecked(False)
                    radio.setEnabled(False)

                radio.style().unpolish(radio)
                radio.style().polish(radio)
                column_layout.addWidget(radio)

            column_layout.addStretch(1)
            self.bucket_columns_layout.addWidget(column_widget, 1)
            if col < FIXED_BUCKET_COLUMNS - 1:
                separator = QFrame()
                separator.setObjectName("bucketSeparator")
                separator.setFrameShape(QFrame.VLine)
                separator.setFrameShadow(QFrame.Plain)
                self.bucket_columns_layout.addWidget(separator)

    def _set_bucket_button_selection(self) -> None:
        for bucket_key, button in self.bucket_button_map.items():
            button.blockSignals(True)
            button.setChecked(bucket_key == self.current_bucket_key)
            button.blockSignals(False)
            button.style().unpolish(button)
            button.style().polish(button)

    def _bucket_window_signature(self, bucket_defs: Optional[List[BucketDef]] = None) -> Tuple[Tuple[str, int], ...]:
        active_defs = bucket_defs if bucket_defs is not None else self.bucket_defs
        return tuple((item.bucket_key, int(item.capacity)) for item in active_defs)

    def _is_bucket_window_stale(self) -> bool:
        if not self.current_source_paths:
            return False
        try:
            bucket_caps = expand_bucket_caps_for_accounts(parse_bucket_config(self.config.bucket_config_json))
            latest_defs = build_window_bucket_defs(get_current_cn_datetime(), bucket_caps)
        except Exception:
            return False
        return self._bucket_window_signature(latest_defs) != self._bucket_window_signature()

    def _refresh_bucket_window_for_new_actions(self, notify_text: str = "") -> bool:
        if self.template2_generated_at_utc is not None:
            return True
        if not self._is_bucket_window_stale():
            return True
        if self.current_source_paths:
            self.load_source_files(self.current_source_paths)
        if notify_text:
            self.status_bar.showMessage(notify_text)
        return False


    def _current_bucket_summary(self) -> str:
        bucket = self.get_current_bucket_def()
        if bucket is None:
            return "无可用桶"
        count = len(self.get_current_bucket_entries())
        return f"当前桶：{bucket.target_date_str} | {bucket.bucket_label} | 人数 {count}/{bucket.capacity}"

    def _template2_base_summary(self) -> str:
        if self.template2_generated_at_utc is None:
            return "模板2基准：未生成"
        local_text = self.template2_generated_at_utc.astimezone(CN_TZ).strftime("%Y/%m/%d %H:%M:%S")
        return f"模板2基准：{local_text}"

    def _get_processing_window_bucket_defs(self, bucket_key: Optional[str] = None) -> List[BucketDef]:
        if not self.bucket_defs:
            return []
        target_key = bucket_key or self.current_bucket_key
        if not target_key:
            return list(self.bucket_defs)
        for idx, item in enumerate(self.bucket_defs):
            if item.bucket_key == target_key:
                return self.bucket_defs[idx:]
        return list(self.bucket_defs)

    def _get_processing_window_end_date(self, bucket_key: Optional[str] = None) -> str:
        window_defs = self._get_processing_window_bucket_defs(bucket_key)
        if not window_defs:
            return ""
        return window_defs[-1].target_date_str

    def _processing_window_summary(self) -> str:
        window_defs = self._get_processing_window_bucket_defs()
        if not window_defs:
            return "处理空间：未分配"
        start_bucket = window_defs[0]
        end_bucket = window_defs[-1]
        return (
            f"处理空间：{start_bucket.target_date_str} {start_bucket.mode}"
            f" → {end_bucket.target_date_str} {end_bucket.mode} | {len(window_defs)}桶"
        )

    def _update_ui_state(self) -> None:
        self.summary_label.setText(
            f"公司{self.summary.company_count}|数据{self.summary.raw_row_count}|有效{self.summary.valid_contact_count}|去重{self.summary.merged_count}|已分配{self.summary.assigned_count}|未分配{self.summary.unassigned_count}|异常{len(self.exceptions)}"
        )
        self._set_state_badge()
        self._update_file_info_labels()
        self._build_bucket_buttons()
        self._set_bucket_button_selection()
        self.bucket_info_label.setText(self._current_bucket_summary())
        self.processing_window_label.setText(self._processing_window_summary())
        self.template2_base_label.setText(self._template2_base_summary())
        current_entries = self.get_current_bucket_entries()
        self.template1_preview.setPlainText(render_template1_preview(current_entries))
        self.template3_preview.setPlainText(render_preview_lines(self.template3_lines))
        self.unassigned_preview.setPlainText(render_unassigned_preview(self.unassigned_items))
        self.exception_preview.setPlainText(render_exception_preview(self.exceptions))
        self.template1_section.set_meta(f"{len(current_entries)} 行")
        self.template2_section.set_meta(f"{count_non_empty_lines(self.template2_input.toPlainText())} 行")
        self.template3_section.set_meta(f"{len(self.template3_lines)} 行")
        self.unassigned_section.set_meta(f"{len(self.unassigned_items)} 条")
        self.exception_section.set_meta(f"{len(self.exceptions)} 条")
        has_bucket_entries = bool(current_entries)
        self.copy_template1_btn.setEnabled(has_bucket_entries)
        self.copy_template2_input_btn.setEnabled(has_bucket_entries)
        self.parse_template2_btn.setEnabled(has_bucket_entries)
        self.clear_template2_btn.setEnabled(bool(self.template2_input.toPlainText().strip()))
        self.copy_template3_btn.setEnabled(bool(self.template3_lines))
        self.export_btn.setEnabled(bool(self.assigned_entries))
        self.copy_company_log_btn.setEnabled(bool(self.company_batches))
        contact_rows_ready = bool(self.template3_lines) and bool(current_entries)
        self.copy_contact_log_btn.setEnabled(contact_rows_ready)
        self._refresh_button_labels()

    def _refresh_button_labels(self) -> None:
        for button in [
            self.copy_template1_btn,
            self.copy_template2_input_btn,
            self.parse_template2_btn,
            self.clear_template2_btn,
            self.copy_template3_btn,
            self.export_btn,
            self.copy_company_log_btn,
            self.copy_contact_log_btn,
        ]:
            base_text = str(button.property("baseText") or button.text().replace(" 禁用", ""))
            button.setText(base_text if button.isEnabled() else f"{base_text} 禁用")

    def _collect_bucket_entries(self, bucket_key: str) -> List[AssignedEntry]:
        if not bucket_key:
            return []
        return [item for item in self.assigned_entries if item.bucket_key == bucket_key]

    def _detect_bucket_key_from_template2_records(self, records: List[Template2Record]) -> Tuple[Optional[str], Dict[str, int]]:
        bucket_counts: Dict[str, int] = {}
        exact_map, normalized_map = build_entry_lookup_maps(self.assigned_entries)
        for record in records:
            entry = resolve_entry_by_row_id(record.row_id, exact_map, normalized_map)
            if entry is None:
                continue
            bucket_counts[entry.bucket_key] = bucket_counts.get(entry.bucket_key, 0) + 1
        if len(bucket_counts) == 1:
            return next(iter(bucket_counts.keys())), bucket_counts
        if len(bucket_counts) > 1:
            return None, bucket_counts
        if self.template2_source_bucket_key and any(item.bucket_key == self.template2_source_bucket_key for item in self.bucket_defs):
            return self.template2_source_bucket_key, bucket_counts
        if self.current_bucket_key and any(item.bucket_key == self.current_bucket_key for item in self.bucket_defs):
            return self.current_bucket_key, bucket_counts
        return None, bucket_counts

    def _summarize_exception_categories(self, exceptions: List[ExceptionItem], limit: int = 4) -> str:
        if not exceptions:
            return "无"
        counts: Dict[str, int] = {}
        for item in exceptions:
            counts[item.category] = counts.get(item.category, 0) + 1
        ranked = sorted(counts.items(), key=lambda x: (-x[1], x[0]))
        return "，".join(f"{name}{count}" for name, count in ranked[:limit])

    def _show_parse_result_feedback(self, success_count: int, current_parse_exceptions: List[ExceptionItem], bucket: Optional[BucketDef]) -> None:
        self.template3_section.set_open(True)
        self.exception_section.set_open(bool(self.exceptions))
        self.unassigned_section.set_open(False)
        bucket_text = bucket.bucket_label if bucket is not None else "未知桶"
        summary = f"解析完成：{bucket_text}，模板3 {success_count} 行，异常 {len(current_parse_exceptions)} 条"
        self.status_bar.showMessage(summary)
        if success_count > 0:
            QMessageBox.information(self, "解析完成", summary + f"\n主要异常：{self._summarize_exception_categories(current_parse_exceptions)}")
        else:
            QMessageBox.warning(self, "解析失败", summary + f"\n主要异常：{self._summarize_exception_categories(current_parse_exceptions)}")

    def _on_template2_text_changed(self) -> None:
        if self.template2_input.toPlainText().strip() and self.get_current_bucket_entries() and self.current_state_text != "模板3已生成":
            self.current_state_text = "模板2已粘贴"
            self._update_ui_state()
        else:
            self.clear_template2_btn.setEnabled(bool(self.template2_input.toPlainText().strip()))
            self.template2_section.set_meta(f"{count_non_empty_lines(self.template2_input.toPlainText())} 行")

    def _reset_current_bucket_results(self, clear_input: bool = True) -> None:
        if clear_input:
            self.template2_input.blockSignals(True)
            self.template2_input.setPlainText("")
            self.template2_input.blockSignals(False)
        self.template2_records = {}
        self.template2_generated_at_utc = None
        self.template2_source_bucket_key = ""
        self.template3_lines = []
        self.template3_exports = []
        self.template3_result_map = {}
        self.current_deferred_export_paths = []

    def _on_bucket_button_clicked(self) -> None:
        button = self.sender()
        if not isinstance(button, QRadioButton):
            return
        if not self._refresh_bucket_window_for_new_actions("分桶窗口已按当前时间刷新，请重新选择当前桶"):
            return
        new_bucket_key = str(button.property("bucketKey") or "")
        if new_bucket_key != self.current_bucket_key:
            self.current_bucket_key = new_bucket_key
            self._reset_current_bucket_results(clear_input=True)
            self.current_state_text = "已就绪" if self.assigned_entries else "未载入文件"
        self._update_ui_state()
        self.status_bar.showMessage(self._current_bucket_summary())

    def dragEnterEvent(self, event) -> None:  # type: ignore[override]
        if event.mimeData().hasUrls() and any(DropFrame._is_excel(url.toLocalFile()) for url in event.mimeData().urls()):
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event) -> None:  # type: ignore[override]
        paths = [url.toLocalFile() for url in event.mimeData().urls() if DropFrame._is_excel(url.toLocalFile())]
        if paths:
            self.handle_dropped_files(paths)
            event.acceptProposedAction()
        else:
            event.ignore()

    def choose_single_file(self) -> None:
        start_dir = self.state.last_input_dir or str(Path.home())
        file_path, _ = QFileDialog.getOpenFileName(self, "选择单个公司文件", start_dir, "Excel Files (*.xlsx *.xls *.xlsm *.xltx *.xltm)")
        if file_path:
            self.load_source_files([Path(file_path)])

    def choose_multiple_files(self) -> None:
        start_dir = self.state.last_input_dir or str(Path.home())
        file_paths, _ = QFileDialog.getOpenFileNames(self, "选择多个公司文件", start_dir, "Excel Files (*.xlsx *.xls *.xlsm *.xltx *.xltm)")
        if file_paths:
            self.load_source_files([Path(path) for path in file_paths])

    def handle_dropped_files(self, file_paths: List[str]) -> None:
        self.load_source_files([Path(path) for path in file_paths])

    def open_settings(self) -> None:
        dialog = SettingsDialog(self.current_config_path, self.config, self)
        if dialog.exec() == QDialog.Accepted:
            try:
                new_config = self.store.load_config(dialog.selected_config_path)
            except Exception as exc:
                QMessageBox.warning(self, "配置切换失败", str(exc))
                return
            self.current_config_path = dialog.selected_config_path
            self.config = new_config
            self.state.config_path = str(self.current_config_path)
            self.store.save_state(self.state)
            self.current_export_path = None
            self.current_deferred_export_paths = []
            self.index_records = self._load_index_records()
            self._update_file_info_labels()
            self.status_bar.showMessage(f"当前配置已切换：{self.current_config_path.name}")
            if self.current_source_paths:
                self.rebuild_from_current_files()
            else:
                self._rebuild_empty_buckets_from_config()
                self._update_ui_state()
                self.status_bar.showMessage("当前配置已切换：未自动载入旧批次，桶已清空")

    def rebuild_from_current_files(self) -> None:
        if self.current_source_paths:
            self.load_source_files(self.current_source_paths)

    def load_source_files(self, file_paths: List[Path]) -> None:
        try:
            bucket_caps = expand_bucket_caps_for_accounts(parse_bucket_config(self.config.bucket_config_json))
        except Exception as exc:
            QMessageBox.warning(self, "配置错误", f"周桶配置有误：\n{exc}")
            return

        self.status_bar.showMessage("正在清洗并分桶…")
        QApplication.processEvents()

        company_batches: List[CompanyBatch] = []
        exceptions: List[ExceptionItem] = []
        summary = ParseSummary(company_count=len(file_paths))

        for path in file_paths:
            try:
                rows = read_sheet_rows(path)
                raw_contacts, removed_b, raw_row_count, _layout = parse_source_rows(rows)
                merged_contacts = build_merged_contacts(raw_contacts)
                company_name = decide_company_name(merged_contacts, path)
                company_batches.append(
                    CompanyBatch(
                        company_key=str(path.resolve()),
                        company_name=company_name,
                        file_path=path,
                        contacts=merged_contacts,
                        raw_row_count=raw_row_count,
                        valid_contact_count=len(raw_contacts),
                        merged_count=len(merged_contacts),
                        removed_b_count=removed_b,
                    )
                )
                summary.raw_row_count += raw_row_count
                summary.valid_contact_count += len(raw_contacts)
                summary.merged_count += len(merged_contacts)
                summary.removed_b_count += removed_b
            except Exception as exc:
                exceptions.append(ExceptionItem("公司文件异常", path.name, str(exc).strip() or "未知错误"))

        batch_code = generate_batch_code()
        current_cn_dt = get_current_cn_datetime()
        bucket_defs = build_window_bucket_defs(current_cn_dt, bucket_caps)
        assigned_entries, unassigned_items, account_overflow_items = assign_weekly_entries(
            company_batches,
            bucket_defs,
            batch_code,
            self.config.company_daily_limit,
        )
        if account_overflow_items:
            exceptions.extend(account_overflow_items)

        self.current_source_paths = file_paths
        self.current_export_path = None
        self.current_deferred_export_paths = []
        self.current_batch_code = batch_code
        self.current_batch_label = f"{current_cn_dt.strftime('%Y%m%d')}_{batch_code}"
        self.company_batches = company_batches
        self.bucket_defs = bucket_defs
        self.allowed_end_cn_date_str = get_allowed_end_cn_date_str(bucket_defs)
        self.assigned_entries = assigned_entries
        self.unassigned_items = unassigned_items
        self.template2_records = {}
        self.template3_lines = []
        self.template3_exports = []
        self.batch_template2_records = {}
        self.batch_template3_exports = {}
        self.batch_exceptions = list(exceptions)
        self.exceptions = exceptions
        self.summary = summary
        self.summary.assigned_count = len(assigned_entries)
        self.summary.unassigned_count = len(unassigned_items)
        self.state.last_input_dir = str(file_paths[0].parent) if file_paths else self.state.last_input_dir
        self.store.save_state(self.state)

        self.current_bucket_key = self._get_first_non_empty_bucket_key()
        self.current_state_text = "已就绪" if assigned_entries else "无可分配联系人"
        self._reset_current_bucket_results(clear_input=True)
        self._update_ui_state()
        self.status_bar.showMessage(f"已完成分桶：{len(company_batches)} 个公司，已分配 {len(assigned_entries)} 条")

    def get_current_bucket_def(self) -> Optional[BucketDef]:
        return next((item for item in self.bucket_defs if item.bucket_key == self.current_bucket_key), None)

    def get_current_bucket_entries(self) -> List[AssignedEntry]:
        if not self.current_bucket_key:
            return []
        return [item for item in self.assigned_entries if item.bucket_key == self.current_bucket_key]

    def copy_template1(self) -> None:
        if not self._refresh_bucket_window_for_new_actions("分桶窗口已按当前时间刷新，请确认当前桶后再复制模板1"):
            return
        entries = self.get_current_bucket_entries()
        if not entries:
            return
        text = "\n".join(f"{item.row_id}|{item.company}|{item.address}" for item in entries)
        QGuiApplication.clipboard().setText(text)
        self.status_bar.showMessage("模板1已复制")

    def build_current_bucket_prompt(self, generated_at_utc: Optional[datetime] = None) -> str:
        entries = self.get_current_bucket_entries()
        if not entries:
            return ""
        bucket = self.get_current_bucket_def()
        if bucket is None:
            return ""
        prompt_head = self.config.morning_prompt if bucket.mode == "AM" else self.config.afternoon_prompt
        base_utc = (generated_at_utc or datetime.now(timezone.utc)).replace(microsecond=0)
        current_utc = base_utc.isoformat().replace("+00:00", "Z")
        allowed_end_cn_date = self._get_processing_window_end_date(bucket.bucket_key) or bucket.target_date_str
        lines = [f"{item.row_id}|{item.company}|{item.address}" for item in entries]
        return (
            f"{prompt_head}\n\n"
            f"current_utc: {current_utc}\n"
            f"target_cn_date: {bucket.target_date_str}\n"
            f"allowed_end_cn_date: {allowed_end_cn_date}\n"
            f"多行联系人数据：\n" + "\n".join(lines)
        )

    def copy_current_bucket_prompt(self) -> None:
        if not self._refresh_bucket_window_for_new_actions("分桶窗口已按当前时间刷新，请确认当前桶后再复制模板2输入稿"):
            return
        bucket = self.get_current_bucket_def()
        if bucket is None or not self.get_current_bucket_entries():
            return
        generated_at_utc = datetime.now(timezone.utc).replace(microsecond=0)
        text = self.build_current_bucket_prompt(generated_at_utc=generated_at_utc)
        if not text:
            return
        self.template2_generated_at_utc = generated_at_utc
        self.template2_source_bucket_key = bucket.bucket_key
        self.current_state_text = "模板2输入稿已生成"
        self._update_ui_state()
        QGuiApplication.clipboard().setText(text)
        self.status_bar.showMessage("模板2输入稿已复制，模板2基准时间已冻结")

    def clear_template2(self) -> None:
        self._reset_current_bucket_results(clear_input=True)
        self.current_state_text = "已就绪" if self.assigned_entries else "未载入文件"
        self._update_ui_state()
        self.status_bar.showMessage("模板2已清空")

    def generate_template3(self) -> None:
        try:
            if not self.assigned_entries:
                QMessageBox.information(self, "提示", "当前没有已分配联系人，无法解析模板2。")
                return
            if self.template2_generated_at_utc is None:
                QMessageBox.warning(self, "缺少模板2基准", "请先点击“复制模板2输入稿”，冻结模板2基准时间后再解析。")
                return
            text = self.template2_input.toPlainText().strip()
            if not text:
                QMessageBox.warning(self, "模板2为空", "请先粘贴 AI 返回的模板2结果，再点击解析。")
                return

            parsed_records, parse_exceptions = parse_template2_text(text)
            self.template2_records = {item.row_id: item for item in parsed_records}
            self.template3_lines = []
            self.template3_exports = []
            self.template3_result_map = {}

            kept_exceptions = [item for item in self.exceptions if item.category == "公司文件异常"]
            self.exceptions = kept_exceptions + parse_exceptions

            detected_bucket_key, bucket_counts = self._detect_bucket_key_from_template2_records(parsed_records)
            if bucket_counts and len(bucket_counts) > 1:
                bucket_map = {item.bucket_key: item for item in self.bucket_defs}
                detail = "，".join(
                    f"{bucket_map[key].bucket_label if key in bucket_map else key}:{count}" for key, count in sorted(bucket_counts.items(), key=lambda x: (-x[1], x[0]))
                )
                mixed_exception = ExceptionItem("混合桶row_id", "", f"模板2同时包含多个桶：{detail}")
                self.exceptions.append(mixed_exception)
                self.current_state_text = "异常"
                self._update_ui_state()
                self.exception_section.set_open(True)
                self.status_bar.showMessage("解析失败：模板2包含多个桶")
                QMessageBox.warning(self, "解析失败", f"模板2同时包含多个桶，无法解析。\n{detail}")
                return

            if detected_bucket_key is None:
                current_parse_exceptions = [item for item in self.exceptions if item.category != "公司文件异常"]
                extra_hint = ""
                if parsed_records:
                    unmatched = sum(1 for item in parsed_records if normalize_generated_row_id(item.row_id))
                    extra_hint = f"\n已解析 {len(parsed_records)} 行，但没有一行能匹配到当前批次分桶。"
                    if unmatched:
                        extra_hint += "\n请优先检查：是否在拿到模板2后又重新分桶/重新导入，导致批次码变化。"
                QMessageBox.warning(self, "解析失败", "未能从模板2中识别所属桶，请确认 row_id 是否来自当前程序生成的模板2输入稿。" + extra_hint)
                self.current_state_text = "异常"
                self._update_ui_state()
                self._show_parse_result_feedback(0, current_parse_exceptions, None)
                return

            if detected_bucket_key != self.current_bucket_key:
                self.current_bucket_key = detected_bucket_key

            bucket = next((item for item in self.bucket_defs if item.bucket_key == detected_bucket_key), None)
            target_entries = self._collect_bucket_entries(detected_bucket_key)
            if not target_entries or bucket is None:
                QMessageBox.warning(self, "解析失败", "识别到的目标桶当前没有已分配联系人。")
                self.current_state_text = "异常"
                self._update_ui_state()
                return

            entry_map = {item.row_id: item for item in target_entries}
            all_entry_map = {item.row_id: item for item in self.assigned_entries}
            entry_map_norm_exact, entry_map_norm = build_entry_lookup_maps(target_entries)
            all_entry_norm_exact, all_entry_norm = build_entry_lookup_maps(self.assigned_entries)
            template2_base_cn = self.template2_generated_at_utc.astimezone(CN_TZ)
            allowed_end_cn_date_str = self._get_processing_window_end_date(detected_bucket_key)
            allowed_end_date = parse_cn_date(allowed_end_cn_date_str).date() if allowed_end_cn_date_str else None
            accepted_candidates: List[Tuple[datetime, AssignedEntry, Template2Record]] = []

            for original_row_id, record in self.template2_records.items():
                entry = resolve_entry_by_row_id(original_row_id, entry_map_norm_exact, entry_map_norm)
                if entry is None:
                    global_entry = resolve_entry_by_row_id(original_row_id, all_entry_norm_exact, all_entry_norm)
                    if global_entry is None:
                        self.exceptions.append(ExceptionItem("未知row_id", original_row_id, "模板2中的 row_id 不存在于当前批次分桶结果"))
                    else:
                        other_bucket = next((item for item in self.bucket_defs if item.bucket_key == global_entry.bucket_key), None)
                        other_label = other_bucket.bucket_label if other_bucket is not None else global_entry.bucket_key
                        self.exceptions.append(ExceptionItem("row_id属于其他桶", original_row_id, f"该 row_id 属于 {other_label}，不是当前解析桶"))
                    continue
                if record.no_valid_time:
                    self.exceptions.append(ExceptionItem("NO_VALID_TIME", original_row_id, "AI 未给出可用时间"))
                    continue
                if parse_cn_date(record.date_str).date() < parse_cn_date(entry.target_date_str).date():
                    self.exceptions.append(ExceptionItem("早于目标日期", original_row_id, f"应晚于或等于 {entry.target_date_str}，实际为 {record.date_str}"))
                    continue
                if allowed_end_date is not None and parse_cn_date(record.date_str).date() > allowed_end_date:
                    self.exceptions.append(ExceptionItem("超出允许范围", original_row_id, f"不得晚于 {allowed_end_cn_date_str}，实际为 {record.date_str}"))
                    continue
                base_dt = datetime.strptime(f"{record.date_str} {record.time_str}", "%Y/%m/%d %H:%M").replace(tzinfo=CN_TZ)
                if base_dt <= template2_base_cn:
                    self.exceptions.append(ExceptionItem("时间倒挂", original_row_id, f"计划营销时间 {record.date_str} {record.time_str} 不晚于模板2基准时间"))
                    continue
                accepted_candidates.append((base_dt, entry, record))

            target_row_keys = {item.row_id for item in target_entries}
            target_row_norm_keys = {normalize_generated_row_id(item.row_id) for item in target_entries if normalize_generated_row_id(item.row_id)}
            returned_norm_keys = {normalize_generated_row_id(row_id) for row_id in self.template2_records.keys() if normalize_generated_row_id(row_id)}
            for item in target_entries:
                if item.row_id in self.template2_records:
                    continue
                norm_key = normalize_generated_row_id(item.row_id)
                if norm_key and norm_key in returned_norm_keys:
                    continue
                self.exceptions.append(ExceptionItem("缺少row_id", item.row_id, "模板2未返回该 row_id"))

            accepted_candidates.sort(key=lambda item: (item[0], item[1].row_id))
            grouped: Dict[Tuple[str, str], List[Tuple[datetime, AssignedEntry]]] = {}

            for base_dt, entry, record in accepted_candidates:
                grouped.setdefault((record.date_str, record.time_str), []).append((base_dt, entry))

            final_items: List[Tuple[datetime, str, Template3ExportRow, AssignedEntry, str, str]] = []
            for (date_str, time_str), grouped_entries in grouped.items():
                base_dt = datetime.strptime(f"{date_str} {time_str}", "%Y/%m/%d %H:%M").replace(tzinfo=CN_TZ)
                offsets = build_safe_offsets(base_dt, len(grouped_entries), template2_base_cn)
                if not offsets:
                    for _, entry in grouped_entries:
                        self.exceptions.append(ExceptionItem("无可用分散时间", entry.row_id, f"{date_str} {time_str} 在分散后无法保持同日且晚于模板2基准时间"))
                    continue
                for (entry_base_dt, entry), offset in zip(sorted(grouped_entries, key=lambda x: x[1].row_id), offsets):
                    final_dt_cn = entry_base_dt + timedelta(minutes=offset)
                    final_date = final_dt_cn.strftime("%Y/%m/%d")
                    final_time = final_dt_cn.strftime("%H:%M")
                    final_line = f"{entry.email_str}|{entry.name}|{final_date}|{final_time}"
                    final_items.append(
                        (
                            final_dt_cn.replace(tzinfo=None),
                            final_line,
                            Template3ExportRow(company=entry.company, name=entry.name, date_str=final_date, time_str=final_time, emails=split_emails(entry.email_str)),
                            entry,
                            final_date,
                            final_time,
                        )
                    )

            final_items.sort(key=lambda x: (x[0], x[1]))
            self.template3_lines = [line for _, line, _, _, _, _ in final_items]
            self.template3_exports = [row for _, _, row, _, _, _ in final_items]
            self.template3_result_map = {
                entry.row_id: (final_date, final_time, line)
                for _, line, _, entry, final_date, final_time in final_items
            }

            bucket_row_ids = {item.row_id for item in target_entries}
            for row_id in bucket_row_ids:
                self.batch_template2_records.pop(row_id, None)
                self.batch_template3_exports.pop(row_id, None)
            for row_id, record in self.template2_records.items():
                self.batch_template2_records[row_id] = record
            for _, _, export_row, entry, _, _ in final_items:
                self.batch_template3_exports[entry.row_id] = export_row

            current_parse_exceptions = [item for item in self.exceptions if item.category != "公司文件异常"]
            self.batch_exceptions = [item for item in self.batch_exceptions if item.category == "公司文件异常"] + current_parse_exceptions
            if self.template3_lines:
                self.current_state_text = "模板3已生成"
            elif current_parse_exceptions:
                self.current_state_text = "异常"
            else:
                self.current_state_text = "模板2已粘贴"

            self._update_ui_state()
            self._show_parse_result_feedback(len(self.template3_lines), current_parse_exceptions, bucket)
        except Exception as exc:
            self.template3_lines = []
            self.template3_exports = []
            self.exceptions.append(ExceptionItem("程序异常", "", str(exc)))
            self.batch_exceptions = [item for item in self.batch_exceptions if item.category == "公司文件异常"] + [item for item in self.exceptions if item.category != "公司文件异常"]
            self.current_state_text = "异常"
            self._update_ui_state()
            self.template3_section.set_open(True)
            self.exception_section.set_open(True)
            self.status_bar.showMessage("解析失败：程序异常")
            QMessageBox.critical(self, "解析失败", f"解析模板2时发生程序异常：\n{exc}")


    def _copy_tsv_rows(self, rows: List[List[str]], success_message: str) -> None:
        if not rows:
            return
        text = "\n".join("\t".join(clean_tsv_cell(value) for value in row) for row in rows)
        QGuiApplication.clipboard().setText(text)
        self.status_bar.showMessage(success_message)

    def get_company_log_rows(self) -> List[List[str]]:
        rows: List[List[str]] = []
        process_time = format_log_timestamp(datetime.now())
        deferred_company_keys = {item.company_key for item in self.unassigned_items}
        for company in self.company_batches:
            company_code = make_company_code(company.company_name or company.file_path.stem)
            status_text = "有顺延" if company.company_key in deferred_company_keys else "队列中"
            rows.append([
                company_code,
                company.file_path.name,
                str(company.file_path),
                status_text,
                process_time,
            ])
        return rows

    def get_contact_log_rows(self) -> List[List[str]]:
        rows: List[List[str]] = []
        current_entries = self.get_current_bucket_entries()
        sortable_items: List[Tuple[Tuple[object, ...], AssignedEntry, str, str, str]] = []
        for entry in current_entries:
            plan_date, plan_time, template3_line = self.template3_result_map.get(entry.row_id, ("", "", ""))
            if not template3_line:
                continue
            company_code = make_company_code(entry.company)
            contact_id = make_contact_id(company_code, entry.linkedin, entry.email_str, entry.name, entry.title)
            plan_text = f"{plan_date} {plan_time}" if plan_date and plan_time else ""
            sort_key = (
                entry.marketing_account_global_seq or 10 ** 9,
                ACCOUNT_DISPLAY_MAP.get(entry.marketing_account, ACCOUNT_COUNT + 1),
                entry.marketing_account_day_seq or 10 ** 9,
                plan_text,
                entry.name.lower(),
                contact_id,
            )
            sortable_items.append((sort_key, entry, company_code, contact_id, plan_text))
        sortable_items.sort(key=lambda x: x[0])
        for _, entry, company_code, contact_id, plan_text in sortable_items:
            _, _, template3_line = self.template3_result_map.get(entry.row_id, ("", "", ""))
            rows.append([
                entry.bucket_key,
                company_code,
                contact_id,
                entry.name,
                entry.title,
                "队列中",
                self.current_batch_label,
                plan_text,
                "",
                entry.source_file,
                entry.email_str,
                template3_line,
                entry.marketing_account,
            ])
        return rows

    def copy_company_log_rows(self) -> None:
        rows = self.get_company_log_rows()
        self._copy_tsv_rows(rows, f"公司日志已复制，共 {len(rows)} 行")

    def copy_contact_log_rows(self) -> None:
        rows = self.get_contact_log_rows()
        self._copy_tsv_rows(rows, f"联系人日志已复制，共 {len(rows)} 行")

    def copy_template3(self) -> None:
        if not self.template3_lines:
            return
        QGuiApplication.clipboard().setText("\n".join(self.template3_lines))
        self.status_bar.showMessage("模板3已复制")

    def export_current_workbook(self) -> None:
        if not self._refresh_bucket_window_for_new_actions("分桶窗口已按当前时间刷新，请确认当前桶后再导出"):
            return
        self._export_current_workbook(show_dialog=True)

    def _export_current_workbook(self, show_dialog: bool) -> Optional[Path]:
        if not self.assigned_entries:
            if show_dialog:
                QMessageBox.information(self, "提示", "当前没有可导出的分配结果。")
            return None
        try:
            export_path = self._ensure_export_path()
            write_export_workbook(
                export_path,
                self.company_batches,
                self.assigned_entries,
                self.bucket_defs,
                self.batch_template2_records,
                self.batch_template3_exports,
                self.unassigned_items,
                self.batch_exceptions,
            )
            source_dir = self.get_effective_source_dir()
            deferred_root = (source_dir / "顺延待处理") if source_dir is not None else (Path(self.config.output_dir).expanduser() / "顺延待处理")
            deferred_map = export_deferred_workbooks(
                deferred_root,
                self.company_batches,
                self.unassigned_items,
            )
            deferred_paths = list(deferred_map.values())
            archive_map = archive_processed_source_files(
                self.current_source_paths,
                source_dir,
                get_current_cn_datetime().strftime("%Y-%m-%d"),
            )
            self.current_source_paths = [archive_map.get(str(path.resolve()), path) for path in self.current_source_paths]
            self.current_deferred_export_paths = deferred_paths
            self.current_export_path = export_path
            self.index_records = update_index_after_export(
                self.index_records,
                self.company_batches,
                self.assigned_entries,
                self.batch_template3_exports,
                self.unassigned_items,
                export_path,
                archive_map,
                deferred_map,
                self.bucket_defs,
            )
            self._save_index_records()
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            self.current_state_text = "异常"
            self._update_ui_state()
            self.status_bar.showMessage("导出失败")
            return None
        self.current_state_text = "已导出"
        self._update_ui_state()
        status_text = f"已导出：{export_path}"
        if deferred_paths:
            status_text += f" | 顺延待处理 {len(deferred_paths)} 份"
        self.status_bar.showMessage(status_text)
        if show_dialog:
            message = f"已导出结果包：\n{export_path}"
            if deferred_paths:
                preview_lines = [str(path) for path in deferred_paths[:8]]
                if len(deferred_paths) > 8:
                    preview_lines.append(f"… 共 {len(deferred_paths)} 份")
                message += "\n\n已更新顺延待处理：\n" + "\n".join(preview_lines)
            QMessageBox.information(self, "导出完成", message)
        return export_path


    def _ensure_export_path(self) -> Path:
        output_dir = Path(self.config.output_dir).expanduser()
        output_dir.mkdir(parents=True, exist_ok=True)
        if self.current_export_path is not None:
            return self.current_export_path
        current_cn = get_current_cn_datetime().strftime("%Y-%m-%d")
        company_count = len(self.company_batches)
        total_people = sum(item.merged_count for item in self.company_batches)
        base_name = f"{current_cn} {company_count}家共{total_people}人"
        export_path = output_dir / f"{base_name}.xlsx"
        self.current_export_path = export_path
        return export_path



def clean_tsv_cell(value: object) -> str:
    text = str(value or "")
    text = text.replace("\t", " ").replace("\r", " ").replace("\n", " ")
    return text.strip()


def format_log_timestamp(dt_value: datetime) -> str:
    return dt_value.strftime("%Y-%m-%d %H:%M:%S")


def stable_text_code(text: str, width: int = 6) -> str:
    raw = hashlib.md5(str(text or "").strip().lower().encode("utf-8")).hexdigest()
    value = int(raw[:12], 16) % (36 ** width)
    return base36(value).upper().zfill(width)


def make_company_code(company_name: str) -> str:
    base_text = norm(company_name) or "UNKNOWN_COMPANY"
    return f"C{stable_text_code(base_text, 6)}"


def make_contact_id(company_code: str, linkedin: str, email_str: str, name: str, title: str) -> str:
    key = norm(linkedin).lower() or norm(email_str).lower() or f"{norm(name).lower()}|{norm(title).lower()}"
    return f"{company_code}-{stable_text_code(company_code + '|' + key, 6)}"


def safe_bucket_object_name(bucket_key: str) -> str:
    text = str(bucket_key or "").strip()
    text = re.sub(r"[^0-9A-Za-z]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "bucket"


def bucket_slot_absolute_index(column_no: int, row_no: int) -> int:
    return (column_no - 1) * FIXED_BUCKET_ROWS + row_no


def build_rpa_bucket_name(x: int, y: int, state: str) -> str:
    return RPA_BUCKET_NAME_TEMPLATE.format(x=x, y=y, state=state)


def build_rpa_bucket_automation_id(x: int, y: int, state: str) -> str:
    return RPA_BUCKET_AUTOMATION_ID_TEMPLATE.format(x=x, y=y, state=state)


def build_rpa_bucket_xpath(x: int, y: int, state: str) -> str:
    return RPA_BUCKET_XPATH_TEMPLATE.format(x=x, y=y, state=state)


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm_text(value: object) -> str:
    return str(value or "").strip()


def is_email(value: object) -> bool:
    return bool(re.search(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", norm_text(value), flags=re.I))


def split_emails(email_str: str) -> List[str]:
    return [part.strip() for part in str(email_str or "").split(";") if part.strip()]


def count_non_empty_lines(text: str) -> int:
    return sum(1 for line in text.splitlines() if line.strip())


def parse_cn_date(text: str) -> datetime:
    return datetime.strptime(text, "%Y/%m/%d")


def get_current_cn_datetime() -> datetime:
    return datetime.now(timezone.utc).astimezone(CN_TZ)


def suggest_monday_str() -> str:
    today = get_current_cn_datetime().date()
    monday = today - timedelta(days=today.weekday())
    return monday.strftime("%Y/%m/%d")


def split_daily_capacity(total_capacity: int) -> Dict[str, int]:
    total_value = max(0, int(total_capacity))
    am_value = (total_value + 1) // 2
    pm_value = total_value // 2
    return {"AM": am_value, "PM": pm_value}


def parse_bucket_config(text: str) -> Dict[str, Dict[str, int]]:
    payload = json.loads(text)
    if not isinstance(payload, dict):
        raise ValueError("根对象必须为对象")
    result: Dict[str, Dict[str, int]] = {}
    for day in WEEKDAY_ORDER:
        if day not in payload:
            raise ValueError(f"缺少 {day} 配置")
        day_value = payload[day]
        if isinstance(day_value, dict):
            am_raw = int(day_value.get("AM", 0))
            pm_raw = int(day_value.get("PM", 0))
            if am_raw < 0 or pm_raw < 0:
                raise ValueError(f"{day} 配置不能为负数")
            result[day] = {"AM": am_raw, "PM": pm_raw}
        else:
            total_value = int(day_value or 0)
            if total_value < 0:
                raise ValueError(f"{day} 日容量不能为负数")
            result[day] = split_daily_capacity(total_value)
    return result


def get_daily_bucket_capacity_map(text: str) -> Dict[str, int]:
    bucket_map = parse_bucket_config(text)
    return {day: int(bucket_map.get(day, {}).get("AM", 0)) + int(bucket_map.get(day, {}).get("PM", 0)) for day in WEEKDAY_ORDER}


def expand_bucket_caps_for_accounts(bucket_caps: Dict[str, Dict[str, int]], account_count: int = ACCOUNT_COUNT) -> Dict[str, Dict[str, int]]:
    multiplier = max(1, int(account_count))
    return {
        day: {
            "AM": int(bucket_caps.get(day, {}).get("AM", 0)) * multiplier,
            "PM": int(bucket_caps.get(day, {}).get("PM", 0)) * multiplier,
        }
        for day in WEEKDAY_ORDER
    }


def next_workday(day_value) -> object:
    current = day_value
    while current.weekday() >= 5:
        current += timedelta(days=1)
    return current


def get_first_available_bucket_slot(now_dt: datetime) -> Tuple[datetime, str]:
    current_date = now_dt.date()
    current_time = now_dt.time()
    am_cutoff = datetime.strptime(MODE_CUTOFF_TIME["AM"], "%H:%M").time()
    pm_cutoff = datetime.strptime(MODE_CUTOFF_TIME["PM"], "%H:%M").time()
    if current_date.weekday() >= 5:
        next_date = next_workday(current_date + timedelta(days=1))
        return datetime.combine(next_date, datetime.min.time()), "AM"
    if current_time < am_cutoff:
        return datetime.combine(current_date, datetime.min.time()), "AM"
    if current_time < pm_cutoff:
        return datetime.combine(current_date, datetime.min.time()), "PM"
    next_date = next_workday(current_date + timedelta(days=1))
    return datetime.combine(next_date, datetime.min.time()), "AM"


def advance_bucket_slot(slot_date: datetime, slot_mode: str) -> Tuple[datetime, str]:
    if slot_mode == "AM":
        return slot_date, "PM"
    next_date = next_workday((slot_date + timedelta(days=1)).date())
    return datetime.combine(next_date, datetime.min.time()), "AM"


def get_next_week_friday(current_date) -> object:
    current_monday = current_date - timedelta(days=current_date.weekday())
    return current_monday + timedelta(days=11)


def build_window_bucket_defs(now_dt: datetime, bucket_caps: Dict[str, Dict[str, int]]) -> List[BucketDef]:
    active_slot_count = sum(1 for day in WEEKDAY_ORDER for mode in ("AM", "PM") if int(bucket_caps[day][mode]) > 0)
    if active_slot_count <= 0:
        return []
    buckets: List[BucketDef] = []
    slot_date, slot_mode = get_first_available_bucket_slot(now_dt)
    end_date = get_next_week_friday(now_dt.date())
    while slot_date.date() <= end_date:
        if slot_date.weekday() < 5:
            weekday_key = WEEKDAY_ORDER[slot_date.weekday()]
            capacity = int(bucket_caps[weekday_key][slot_mode])
            if capacity > 0:
                target_date_str = slot_date.strftime("%Y/%m/%d")
                bucket_label = f"{weekday_key}-{slot_mode}"
                buckets.append(
                    BucketDef(
                        bucket_key=f"{target_date_str}|{bucket_label}",
                        bucket_label=bucket_label,
                        weekday_key=weekday_key,
                        weekday_label=WEEKDAY_LABELS[weekday_key],
                        mode=slot_mode,
                        mode_label=MODE_LABELS[slot_mode],
                        target_date_str=target_date_str,
                        capacity=capacity,
                    )
                )
        slot_date, slot_mode = advance_bucket_slot(slot_date, slot_mode)
    return buckets


def get_allowed_end_cn_date_str(bucket_defs: List[BucketDef]) -> str:
    if not bucket_defs:
        return ""
    return max(item.target_date_str for item in bucket_defs)


def decide_company_name(contacts: List[ContactRecord], path: Path) -> str:
    values = [norm(item.company) for item in contacts if norm(item.company)]
    if values:
        counts: Dict[str, int] = {}
        for value in values:
            counts[value] = counts.get(value, 0) + 1
        values.sort(key=lambda x: (counts[x], len(x), x), reverse=True)
        return values[0]
    return path.stem


def base36(number: int) -> str:
    digits = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if number == 0:
        return "0"
    parts = []
    while number:
        number, remainder = divmod(number, 36)
        parts.append(digits[remainder])
    return "".join(reversed(parts))


def generate_batch_code() -> str:
    now = datetime.now(timezone.utc)
    value = int(now.timestamp()) % (36 ** 2)
    return base36(value).upper().zfill(2)[-2:]


def interleave_entries_light(entries: List[AssignedEntry]) -> List[AssignedEntry]:
    if len(entries) <= 2:
        return list(entries)
    queues: Dict[str, deque] = {}
    company_order: Dict[str, int] = {}
    for entry in entries:
        company_key = entry.company_key or entry.company or "__EMPTY__"
        if company_key not in queues:
            queues[company_key] = deque()
            company_order[company_key] = len(company_order)
        queues[company_key].append(entry)
    result: List[AssignedEntry] = []
    last_company_key = ""
    while len(result) < len(entries):
        candidates: List[Tuple[int, int, str]] = []
        for company_key, queue in queues.items():
            if not queue:
                continue
            if company_key == last_company_key:
                continue
            candidates.append((-len(queue), company_order[company_key], company_key))
        if not candidates:
            for company_key, queue in queues.items():
                if queue:
                    candidates.append((-len(queue), company_order[company_key], company_key))
        candidates.sort()
        chosen_company = candidates[0][2]
        entry = queues[chosen_company].popleft()
        result.append(entry)
        last_company_key = chosen_company
    return result


def build_account_assignment_failure_message(
    date_str: str,
    entry: AssignedEntry,
    account_totals: List[int],
    account_company_totals: List[Dict[str, int]],
    company_daily_limit: int,
) -> str:
    summary_parts: List[str] = []
    for idx, marker_name in enumerate(ACCOUNT_MARKERS):
        company_used = account_company_totals[idx].get(entry.company_key, 0)
        summary_parts.append(f"{marker_name}:{account_totals[idx]}/50,同公司{company_used}/{company_daily_limit}")
    summary_text = "；".join(summary_parts)
    return (
        f"营销账号分配失败：{date_str} | 公司 {entry.company} | 联系人 {entry.name}。\n"
        f"当前联系人已无法继续顺延到后续账号链，请检查：\n"
        f"1. 单账号同公司单日上限是否过低；\n"
        f"2. 当天该公司的联系人是否过于集中；\n"
        f"3. 当天总容量是否足够。\n\n"
        f"账号占用快照：{summary_text}"
    )



def _build_company_queues(entries: List[AssignedEntry]) -> Tuple[Dict[str, deque], Dict[str, int]]:
    ordered_entries = list(interleave_entries_light(entries))
    queues: Dict[str, deque] = {}
    company_order: Dict[str, int] = {}
    for idx, entry in enumerate(ordered_entries):
        queues.setdefault(entry.company_key, deque()).append(entry)
        company_order.setdefault(entry.company_key, idx)
    return queues, company_order


def _pick_company_for_account_block(
    queues: Dict[str, deque],
    company_order: Dict[str, int],
    account_index: int,
    date_str: str,
    account_company_totals: Dict[Tuple[int, str, str], int],
    per_account_company_limit: int,
) -> Optional[str]:
    candidates: List[Tuple[int, int, str]] = []
    for company_key, queue in queues.items():
        if not queue:
            continue
        used_count = account_company_totals.get((account_index, date_str, company_key), 0)
        remaining_company_quota = per_account_company_limit - used_count
        if remaining_company_quota <= 0:
            continue
        candidates.append((-len(queue), company_order.get(company_key, 10 ** 9), company_key))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][2]


def _build_daily_account_targets(total_count: int, active_account_indices: List[int]) -> Dict[int, int]:
    if total_count <= 0 or not active_account_indices:
        return {}
    targets: Dict[int, int] = {idx: 0 for idx in active_account_indices}
    remaining = int(total_count)

    if remaining >= len(active_account_indices):
        for idx in active_account_indices:
            targets[idx] += 1
        remaining -= len(active_account_indices)

    weighted_cycle: List[int] = []
    for idx in active_account_indices:
        marker_name = ACCOUNT_MARKERS[idx]
        weight = max(1, int(ACCOUNT_WEIGHT_MAP.get(marker_name, 1)))
        weighted_cycle.extend([idx] * weight)

    while remaining > 0:
        progressed = False
        for idx in weighted_cycle:
            if remaining <= 0:
                break
            if targets[idx] >= 50:
                continue
            targets[idx] += 1
            remaining -= 1
            progressed = True
        if not progressed:
            break
    return targets



def finalize_assigned_entries(
    assigned_entries: List[AssignedEntry],
    bucket_defs: List[BucketDef],
    batch_code: str,
    company_daily_limit: int,
) -> Tuple[List[AssignedEntry], List[UnassignedItem], List[ExceptionItem]]:
    entries_by_bucket: Dict[str, List[AssignedEntry]] = {}
    for entry in assigned_entries:
        entries_by_bucket.setdefault(entry.bucket_key, []).append(entry)

    finalized_by_bucket: Dict[str, List[AssignedEntry]] = {item.bucket_key: [] for item in bucket_defs}
    account_limited_unassigned: List[UnassignedItem] = []
    account_limit_exceptions: List[ExceptionItem] = []
    per_account_company_limit = max(1, int(company_daily_limit))

    account_day_totals: Dict[Tuple[int, str], int] = {}
    account_company_totals: Dict[Tuple[int, str, str], int] = {}
    account_day_seq: Dict[Tuple[int, str], int] = {}
    global_batch_seq = 0

    ordered_buckets = sorted(
        bucket_defs,
        key=lambda x: (x.target_date_str, 0 if x.mode == "AM" else 1, x.bucket_key),
    )

    buckets_by_date: Dict[str, List[BucketDef]] = {}
    for bucket in ordered_buckets:
        buckets_by_date.setdefault(bucket.target_date_str, []).append(bucket)

    for date_str in sorted(buckets_by_date.keys()):
        day_buckets = buckets_by_date[date_str]
        day_company_counts: Dict[str, int] = {}
        total_day_entries = 0
        for bucket in day_buckets:
            bucket_entries = entries_by_bucket.get(bucket.bucket_key, [])
            total_day_entries += len(bucket_entries)
            for entry in bucket_entries:
                day_company_counts[entry.company_key] = day_company_counts.get(entry.company_key, 0) + 1
        if total_day_entries <= 0:
            continue

        required_by_total = (total_day_entries + 49) // 50
        required_by_company = 1
        if day_company_counts:
            required_by_company = max((count + per_account_company_limit - 1) // per_account_company_limit for count in day_company_counts.values())
        active_account_count = min(ACCOUNT_COUNT, max(required_by_total, required_by_company))
        selected_priority_indices = ACCOUNT_SELECTION_INDICES[:active_account_count]
        active_account_indices = sorted(selected_priority_indices, key=lambda idx: ACCOUNT_DISPLAY_MAP.get(ACCOUNT_MARKERS[idx], 10 ** 9))
        daily_account_targets = _build_daily_account_targets(total_day_entries, selected_priority_indices)
        current_account_pointer = 0

        for bucket in day_buckets:
            queues, company_order = _build_company_queues(entries_by_bucket.get(bucket.bucket_key, []))
            while any(queue for queue in queues.values()) and current_account_pointer < len(active_account_indices):
                current_account_index = active_account_indices[current_account_pointer]
                current_target = daily_account_targets.get(current_account_index, 0)
                current_used = account_day_totals.get((current_account_index, date_str), 0)
                target_left = current_target - current_used
                if target_left <= 0:
                    current_account_pointer += 1
                    continue
                account_capacity_left = min(50 - current_used, target_left)
                if account_capacity_left <= 0:
                    current_account_pointer += 1
                    continue

                chosen_company = _pick_company_for_account_block(
                    queues,
                    company_order,
                    current_account_index,
                    date_str,
                    account_company_totals,
                    per_account_company_limit,
                )
                if chosen_company is None:
                    current_account_pointer += 1
                    continue

                used_count = account_company_totals.get((current_account_index, date_str, chosen_company), 0)
                remaining_company_quota = per_account_company_limit - used_count
                assign_count = min(account_capacity_left, remaining_company_quota, len(queues[chosen_company]))
                if assign_count <= 0:
                    current_account_pointer += 1
                    continue

                for _ in range(assign_count):
                    entry = queues[chosen_company].popleft()
                    entry.marketing_account = ACCOUNT_MARKERS[current_account_index]
                    entry.marketing_account_index = current_account_index
                    account_day_totals[(current_account_index, date_str)] = account_day_totals.get((current_account_index, date_str), 0) + 1
                    account_day_seq[(current_account_index, date_str)] = account_day_seq.get((current_account_index, date_str), 0) + 1
                    global_batch_seq += 1
                    entry.marketing_account_day_seq = account_day_seq[(current_account_index, date_str)]
                    entry.marketing_account_global_seq = global_batch_seq
                    account_company_totals[(current_account_index, date_str, entry.company_key)] = (
                        account_company_totals.get((current_account_index, date_str, entry.company_key), 0) + 1
                    )
                    finalized_by_bucket[bucket.bucket_key].append(entry)

            remaining_entries: List[AssignedEntry] = []
            for queue in queues.values():
                while queue:
                    remaining_entries.append(queue.popleft())
            if remaining_entries:
                remaining_entries.sort(key=lambda entry: (company_order.get(entry.company_key, 10 ** 9), entry.name.lower(), entry.source_rows[:1] or [entry.source_file]))
                for entry in remaining_entries:
                    account_limited_unassigned.append(
                        UnassignedItem(
                            company_key=entry.company_key,
                            company=entry.company,
                            source_file=entry.source_file,
                            name=entry.name,
                            title=entry.title,
                            linkedin=entry.linkedin,
                            address=entry.address,
                            email_str=entry.email_str,
                            reason="营销账号容量/同公司上限，已顺延",
                            source_rows=list(entry.source_rows),
                        )
                    )
                    account_limit_exceptions.append(
                        ExceptionItem(
                            "营销账号顺延",
                            entry.row_id,
                            f"{entry.target_date_str} | {entry.company} 因营销账号容量/同公司上限已顺延",
                        )
                    )

    finalized_entries: List[AssignedEntry] = []
    row_seq_by_bucket: Dict[str, int] = {item.bucket_key: 1 for item in bucket_defs}
    ordered_final_buckets = sorted(bucket_defs, key=lambda x: (x.target_date_str, 0 if x.mode == "AM" else 1, x.bucket_key))
    for bucket in ordered_final_buckets:
        bucket_items = finalized_by_bucket.get(bucket.bucket_key, [])
        bucket_items.sort(
            key=lambda entry: (
                entry.marketing_account_global_seq or 10 ** 9,
                ACCOUNT_DISPLAY_MAP.get(entry.marketing_account, ACCOUNT_COUNT + 1),
                entry.marketing_account_day_seq or 10 ** 9,
                norm(entry.company).lower(),
                (entry.source_rows[0] if entry.source_rows else 10 ** 9),
                entry.name.lower(),
            )
        )
        for entry in bucket_items:
            seq = row_seq_by_bucket[bucket.bucket_key]
            row_seq_by_bucket[bucket.bucket_key] += 1
            entry.row_id = f"{batch_code}{compress_cn_date(bucket.target_date_str)}{bucket.mode[0]}{seq:03d}"
            finalized_entries.append(entry)
    return finalized_entries, account_limited_unassigned, account_limit_exceptions

def assign_weekly_entries(
    company_batches: List[CompanyBatch],
    bucket_defs: List[BucketDef],
    batch_code: str,
    company_daily_limit: int,
) -> Tuple[List[AssignedEntry], List[UnassignedItem], List[ExceptionItem]]:
    bucket_remaining = {item.bucket_key: item.capacity for item in bucket_defs}
    company_day_count: Dict[Tuple[str, str], int] = {}
    assigned: List[AssignedEntry] = []
    unassigned: List[UnassignedItem] = []
    overall_company_day_limit = max(1, int(company_daily_limit)) * ACCOUNT_COUNT

    for company in company_batches:
        for contact in company.contacts:
            placed_bucket: Optional[BucketDef] = None
            for bucket in bucket_defs:
                if bucket_remaining[bucket.bucket_key] <= 0:
                    continue
                daily_key = (company.company_key, bucket.target_date_str)
                if company_day_count.get(daily_key, 0) >= overall_company_day_limit:
                    continue
                placed_bucket = bucket
                break
            if placed_bucket is None:
                remaining_buckets = [bucket for bucket in bucket_defs if bucket_remaining[bucket.bucket_key] > 0]
                if not remaining_buckets:
                    reason = "周桶容量不足"
                elif all(company_day_count.get((company.company_key, bucket.target_date_str), 0) >= overall_company_day_limit for bucket in remaining_buckets):
                    reason = f"单公司单日上限（{ACCOUNT_COUNT}账号合计）"
                else:
                    reason = "周桶容量不足"
                unassigned.append(
                    UnassignedItem(
                        company_key=company.company_key,
                        company=company.company_name,
                        source_file=company.file_path.name,
                        name=contact.name,
                        title=contact.title,
                        linkedin=contact.linkedin,
                        address=contact.address,
                        email_str=contact.email_str,
                        reason=reason,
                        source_rows=list(contact.source_rows or [contact.source_row]),
                    )
                )
                continue
            bucket_remaining[placed_bucket.bucket_key] -= 1
            company_day_count[(company.company_key, placed_bucket.target_date_str)] = company_day_count.get((company.company_key, placed_bucket.target_date_str), 0) + 1
            assigned.append(
                AssignedEntry(
                    row_id="",
                    company_key=company.company_key,
                    company=company.company_name,
                    source_file=company.file_path.name,
                    name=contact.name,
                    title=contact.title,
                    linkedin=contact.linkedin,
                    address=contact.address,
                    email_str=contact.email_str,
                    bucket_key=placed_bucket.bucket_key,
                    weekday_label=placed_bucket.weekday_label,
                    mode=placed_bucket.mode,
                    mode_label=placed_bucket.mode_label,
                    target_date_str=placed_bucket.target_date_str,
                    marketing_account="",
                    marketing_account_index=-1,
                    marketing_account_day_seq=0,
                    source_rows=list(contact.source_rows or [contact.source_row]),
                )
            )
    finalized_assigned, account_limited_unassigned, account_limit_exceptions = finalize_assigned_entries(
        assigned,
        bucket_defs,
        batch_code,
        company_daily_limit,
    )
    unassigned.extend(account_limited_unassigned)
    return finalized_assigned, unassigned, account_limit_exceptions

def close_workbook_safely(workbook: object) -> None:
    close_method = getattr(workbook, "close", None)
    if callable(close_method):
        try:
            close_method()
        except Exception:
            pass


def read_sheet_rows(file_path: Path) -> List[List[object]]:
    ext = file_path.suffix.lower()
    if ext == ".xls":
        raise RuntimeError("当前版本仅直接支持 xlsx/xlsm/xltx/xltm；若需读取 xls，请先转换为 xlsx。")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        close_workbook_safely(wb)


def normalize_header_key(value: object) -> str:
    text = norm_text(value).lower()
    text = re.sub(r"[\s\-_:/|（）()【】\[\]·,.]+", "", text)
    return text


def is_email_header_key(key: str) -> bool:
    if not key:
        return False
    return (("邮箱" in key) or ("email" in key)) and ("标识" not in key) and ("marker" not in key)


def is_marker_header_key(key: str) -> bool:
    if not key:
        return False
    return ("标识" in key) or ("marker" in key)


def header_contains_any(key: str, words: Tuple[str, ...]) -> bool:
    return any(word in key for word in words)


def digits_in_key(key: str) -> str:
    match = re.search(r"(\d+)", key)
    return match.group(1) if match else ""


def row_preview_text(row: List[object], limit: int = 12) -> str:
    parts = [norm_text(cell) for cell in row[:limit] if norm_text(cell)]
    if not parts:
        return "（空）"
    if len(parts) < len([cell for cell in row if norm_text(cell)]):
        parts.append("…")
    return " | ".join(parts)


def pad_keys(row: List[str], size: int) -> List[str]:
    return row + [""] * max(0, size - len(row))


def merge_header_rows(primary_row: List[str], secondary_row: List[str]) -> List[str]:
    size = max(len(primary_row), len(secondary_row))
    first = pad_keys(primary_row, size)
    second = pad_keys(secondary_row, size)
    merged: List[str] = []
    for a, b in zip(first, second):
        if a and b:
            merged.append(f"{a}{b}")
        else:
            merged.append(a or b)
    return merged


def base_header_hit_count(key_row: List[str], header_rules: Dict[str, Tuple[str, ...]]) -> int:
    hits = 0
    for aliases in header_rules.values():
        if any(header_contains_any(key, aliases) for key in key_row):
            hits += 1
    return hits


def locate_source_layout(rows: List[List[object]]) -> SourceLayout:
    if not rows:
        raise RuntimeError("Excel 为空，无法读取联系人数据。")
    header_rules = {
        "name": ("姓名", "name", "联系人", "contactname"),
        "title": ("职位", "职业", "title", "jobtitle", "position"),
        "linkedin": ("linkedin", "领英", "linkedln"),
        "company": ("公司", "company", "企业", "companyname", "公司名称"),
        "address": ("地址", "address", "location", "所在地", "地区"),
    }
    best_row_index: Optional[int] = None
    best_score = -1
    best_key_row: List[str] = []
    best_use_second_row = False
    scan_limit = min(len(rows), 8)
    for row_index in range(scan_limit):
        key_row = [normalize_header_key(cell) for cell in rows[row_index]]
        if not any(key_row):
            continue
        next_key_row: List[str] = []
        if row_index + 1 < len(rows):
            next_key_row = [normalize_header_key(cell) for cell in rows[row_index + 1]]
        combined_row = merge_header_rows(key_row, next_key_row)
        base_hits_single = base_header_hit_count(key_row, header_rules)
        email_hits_single = sum(1 for key in key_row if is_email_header_key(key))
        marker_hits_single = sum(1 for key in key_row if is_marker_header_key(key))
        base_hits_combined = base_header_hit_count(combined_row, header_rules)
        email_hits_combined = sum(1 for key in combined_row if is_email_header_key(key))
        marker_hits_combined = sum(1 for key in combined_row if is_marker_header_key(key))
        second_header_tokens = sum(1 for key in next_key_row if is_email_header_key(key) or is_marker_header_key(key))
        use_second_row = second_header_tokens >= 2 and email_hits_combined > email_hits_single
        active_row = combined_row if use_second_row else key_row
        base_hits = base_hits_combined if use_second_row else base_hits_single
        email_hits = email_hits_combined if use_second_row else email_hits_single
        marker_hits = marker_hits_combined if use_second_row else marker_hits_single
        score = base_hits * 10 + email_hits * 4 + marker_hits
        if use_second_row:
            score += 8
        if base_hits >= 2 and email_hits >= 1:
            score += 20
        if score > best_score:
            best_score = score
            best_row_index = row_index
            best_key_row = active_row
            best_use_second_row = use_second_row
    if best_row_index is None or best_score < 20:
        previews = []
        for idx in range(min(4, len(rows))):
            previews.append(f"第{idx + 1}行：{row_preview_text(rows[idx])}")
        raise RuntimeError("未识别到可用表头。\n\n当前文件前几行预览：\n" + "\n".join(previews))

    def find_col(alias_group: Tuple[str, ...]) -> Optional[int]:
        for col_index, key in enumerate(best_key_row):
            if header_contains_any(key, alias_group):
                return col_index
        return None

    name_col = find_col(header_rules["name"])
    title_col = find_col(header_rules["title"])
    linkedin_col = find_col(header_rules["linkedin"])
    company_col = find_col(header_rules["company"])
    address_col = find_col(header_rules["address"])

    email_pairs: List[Tuple[Optional[int], int]] = []
    email_candidates: Dict[str, int] = {}
    marker_candidates: Dict[str, int] = {}
    for col_index, key in enumerate(best_key_row):
        if is_email_header_key(key):
            email_candidates[digits_in_key(key) or str(col_index)] = col_index
        elif is_marker_header_key(key):
            marker_candidates[digits_in_key(key) or str(col_index)] = col_index
    for num_key, email_col in email_candidates.items():
        email_pairs.append((marker_candidates.get(num_key), email_col))
    if not email_pairs:
        raise RuntimeError("未识别到邮箱列。")
    email_pairs.sort(key=lambda x: x[1])
    return SourceLayout(
        header_row_index=best_row_index,
        data_start_index=best_row_index + (2 if best_use_second_row else 1),
        name_col=name_col,
        title_col=title_col,
        linkedin_col=linkedin_col,
        company_col=company_col,
        address_col=address_col,
        email_pairs=email_pairs,
    )


def parse_source_rows(rows: List[List[object]]) -> Tuple[List[ContactRecord], int, int, SourceLayout]:
    layout = locate_source_layout(rows)
    removed_b = 0
    raw_row_count = 0
    contacts: List[ContactRecord] = []

    def get_value(row: List[object], col_index: Optional[int]) -> str:
        if col_index is None or col_index >= len(row):
            return ""
        return norm(row[col_index])

    for idx in range(layout.data_start_index, len(rows)):
        row = rows[idx] if idx < len(rows) else []
        if not any(norm_text(cell) for cell in row):
            continue
        raw_row_count += 1
        name = get_value(row, layout.name_col)
        title = get_value(row, layout.title_col)
        linkedin = get_value(row, layout.linkedin_col)
        company = get_value(row, layout.company_col)
        address = get_value(row, layout.address_col)
        emails: List[str] = []
        seen = set()
        for marker_col, email_col in layout.email_pairs:
            email_value = norm_text(row[email_col] if email_col < len(row) else "").lower()
            marker_value = norm_text(row[marker_col] if marker_col is not None and marker_col < len(row) else "").upper()
            if not email_value or not is_email(email_value):
                continue
            if marker_value.startswith("B"):
                removed_b += 1
                continue
            if email_value not in seen:
                seen.add(email_value)
                emails.append(email_value)
        if not emails:
            continue
        contacts.append(ContactRecord(source_row=idx + 1, company=company, name=name, title=title, linkedin=linkedin, address=address, emails=emails, email_str=";".join(emails), source_rows=[idx + 1]))
    return contacts, removed_b, raw_row_count, layout


def build_merged_contacts(raw_contacts: List[ContactRecord]) -> List[ContactRecord]:
    total = len(raw_contacts)
    if total == 0:
        return []
    parent = list(range(total))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    email_owner: Dict[str, int] = {}
    linkedin_owner: Dict[str, int] = {}
    for idx, item in enumerate(raw_contacts):
        for email in item.emails:
            if email in email_owner:
                union(idx, email_owner[email])
            else:
                email_owner[email] = idx
        linkedin_key = norm(item.linkedin).lower()
        if linkedin_key:
            if linkedin_key in linkedin_owner:
                union(idx, linkedin_owner[linkedin_key])
            else:
                linkedin_owner[linkedin_key] = idx

    groups: Dict[int, List[ContactRecord]] = {}
    for idx, item in enumerate(raw_contacts):
        groups.setdefault(find(idx), []).append(item)

    merged: List[ContactRecord] = []
    for items in groups.values():
        items_sorted = sorted(items, key=lambda x: x.source_row)
        email_seen = set()
        email_list: List[str] = []
        for item in items_sorted:
            for email in item.emails:
                if email not in email_seen:
                    email_seen.add(email)
                    email_list.append(email)

        def pick_longest(attr: str) -> str:
            values = [norm(getattr(item, attr)) for item in items_sorted if norm(getattr(item, attr))]
            values.sort(key=lambda v: (len(v), v), reverse=True)
            return values[0] if values else ""

        merged.append(ContactRecord(source_row=min(item.source_row for item in items_sorted), company=pick_longest("company"), name=pick_longest("name"), title=pick_longest("title"), linkedin=pick_longest("linkedin"), address=pick_longest("address"), emails=email_list, email_str=";".join(email_list), source_rows=sorted({row_no for item in items_sorted for row_no in (item.source_rows or [item.source_row])})))
    merged.sort(key=lambda x: x.source_row)
    return merged


def render_template1_preview(entries: List[AssignedEntry]) -> str:
    if not entries:
        return ""
    lines = [f"{item.row_id}|{item.company}|{item.address}" for item in entries[:8]]
    if len(entries) > 8:
        lines.append(f"… 共 {len(entries)} 行")
    return "\n".join(lines)


def render_preview_lines(lines: List[str]) -> str:
    if not lines:
        return ""
    preview = lines[:8]
    if len(lines) > 8:
        preview.append(f"… 共 {len(lines)} 行")
    return "\n".join(preview)


def render_unassigned_preview(items: List[UnassignedItem]) -> str:
    if not items:
        return ""
    lines = [f"{item.company}|{item.name}|{item.email_str}|{item.reason}" for item in items[:8]]
    if len(items) > 8:
        lines.append(f"… 共 {len(items)} 条")
    return "\n".join(lines)


def render_exception_preview(items: List[ExceptionItem]) -> str:
    if not items:
        return ""
    lines = [f"{item.category}|{item.row_id}|{item.detail}" for item in items[:8]]
    if len(items) > 8:
        lines.append(f"… 共 {len(items)} 条")
    return "\n".join(lines)


def build_date_capacity_map(bucket_defs: List[BucketDef]) -> Dict[str, int]:
    capacity_by_date: Dict[str, int] = {}
    for item in bucket_defs:
        capacity_by_date[item.target_date_str] = capacity_by_date.get(item.target_date_str, 0) + item.capacity
    return capacity_by_date


def build_reserved_capacity_maps(
    assigned_entries: List[AssignedEntry],
    current_bucket_key: str,
) -> Tuple[Dict[str, int], Dict[Tuple[str, str], int]]:
    date_usage: Dict[str, int] = {}
    company_usage: Dict[Tuple[str, str], int] = {}
    for item in assigned_entries:
        if item.bucket_key == current_bucket_key:
            continue
        date_usage[item.target_date_str] = date_usage.get(item.target_date_str, 0) + 1
        company_key = (item.company_key, item.target_date_str)
        company_usage[company_key] = company_usage.get(company_key, 0) + 1
    return date_usage, company_usage


def compress_cn_date(date_str: str) -> str:
    return date_str.replace("/", "")[4:8]


def normalize_generated_row_id(row_id: str) -> str:
    text = str(row_id or "").strip().upper()
    match = GENERATED_ROW_ID_RE.match(text)
    if match:
        return match.group(1)
    return text


def build_entry_lookup_maps(entries: List[AssignedEntry]) -> Tuple[Dict[str, AssignedEntry], Dict[str, AssignedEntry]]:
    exact_map: Dict[str, AssignedEntry] = {}
    normalized_map: Dict[str, AssignedEntry] = {}
    collision_keys = set()
    for entry in entries:
        exact_map[entry.row_id] = entry
        normalized_key = normalize_generated_row_id(entry.row_id)
        if not normalized_key:
            continue
        if normalized_key in normalized_map and normalized_map[normalized_key].row_id != entry.row_id:
            collision_keys.add(normalized_key)
            continue
        normalized_map[normalized_key] = entry
    for key in collision_keys:
        normalized_map.pop(key, None)
    return exact_map, normalized_map


def resolve_entry_by_row_id(row_id: str, exact_map: Dict[str, AssignedEntry], normalized_map: Dict[str, AssignedEntry]) -> Optional[AssignedEntry]:
    direct = exact_map.get(row_id)
    if direct is not None:
        return direct
    return normalized_map.get(normalize_generated_row_id(row_id))


def parse_template2_text(text: str) -> Tuple[List[Template2Record], List[ExceptionItem]]:
    records: List[Template2Record] = []
    exceptions: List[ExceptionItem] = []
    seen = set()
    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) == 2 and parts[1] == "NO_VALID_TIME":
            row_id = parts[0]
            if row_id in seen:
                exceptions.append(ExceptionItem("重复row_id", row_id, f"模板2第 {line_no} 行重复"))
                continue
            seen.add(row_id)
            records.append(Template2Record(row_id=row_id, no_valid_time=True))
            continue
        if len(parts) != 3:
            exceptions.append(ExceptionItem("格式错误", "", f"模板2第 {line_no} 行格式不正确：{line}"))
            continue
        row_id, date_str, time_str = parts
        if row_id in seen:
            exceptions.append(ExceptionItem("重复row_id", row_id, f"模板2第 {line_no} 行重复"))
            continue
        if not TEMPLATE2_TIME_RE.match(f"{date_str}|{time_str}"):
            exceptions.append(ExceptionItem("时间格式错误", row_id, f"模板2第 {line_no} 行时间不合法：{date_str}|{time_str}"))
            continue
        seen.add(row_id)
        records.append(Template2Record(row_id=row_id, date_str=date_str, time_str=time_str))
    return records, exceptions


def select_even_slot_indices(count: int, slot_count: int = 25) -> List[int]:
    if count <= 0 or slot_count <= 0:
        return []
    if slot_count == 1:
        return [0] * count
    max_index = slot_count - 1
    if count == 1:
        return [max_index // 2]
    return [min(max_index, max(0, int(round((max_index * idx) / (count - 1))))) for idx in range(count)]


def build_even_offsets(count: int) -> List[int]:
    if count <= 0:
        return []
    slot_values = [offset for offset in range(-120, 121, 10)]
    if count == 1:
        non_center = [value for value in slot_values if value != 0]
        return [random.choice(non_center)]
    if count <= len(slot_values):
        return [slot_values[idx] for idx in select_even_slot_indices(count, len(slot_values))]
    full_rounds, remainder = divmod(count, len(slot_values))
    slot_counts = [full_rounds] * len(slot_values)
    for idx in select_even_slot_indices(remainder, len(slot_values)):
        slot_counts[idx] += 1
    offsets: List[int] = []
    for slot_value, repeat_count in zip(slot_values, slot_counts):
        offsets.extend([slot_value] * repeat_count)
    return offsets


def build_safe_offsets(base_dt: datetime, count: int, lower_bound_dt: datetime) -> List[int]:
    allowed_offsets = []
    for offset in range(-120, 121, 10):
        final_dt = base_dt + timedelta(minutes=offset)
        if final_dt.date() != base_dt.date():
            continue
        if final_dt <= lower_bound_dt:
            continue
        allowed_offsets.append(offset)
    if not allowed_offsets:
        return []
    if count == 1:
        non_center = [value for value in allowed_offsets if value != 0]
        return [random.choice(non_center or allowed_offsets)]
    if count <= len(allowed_offsets):
        return [allowed_offsets[idx] for idx in select_even_slot_indices(count, len(allowed_offsets))]
    full_rounds, remainder = divmod(count, len(allowed_offsets))
    slot_counts = [full_rounds] * len(allowed_offsets)
    for idx in select_even_slot_indices(remainder, len(allowed_offsets)):
        slot_counts[idx] += 1
    offsets: List[int] = []
    for slot_value, repeat_count in zip(allowed_offsets, slot_counts):
        offsets.extend([slot_value] * repeat_count)
    return offsets


def ceil_to_next_10_minutes(dt_value: datetime) -> datetime:
    dt_value = dt_value.replace(second=0, microsecond=0)
    remainder = dt_value.minute % 10
    if remainder == 0:
        return dt_value + timedelta(minutes=10)
    return dt_value + timedelta(minutes=(10 - remainder))


def build_non_overwriting_path(output_dir: Path, base_name: str, suffix: str) -> Path:
    candidate = output_dir / f"{base_name}{suffix}"
    if not candidate.exists():
        return candidate
    index = 1
    while True:
        candidate = output_dir / f"{base_name}({index}){suffix}"
        if not candidate.exists():
            return candidate
        index += 1



def write_deferred_source_workbook(source_path: Path, kept_source_rows: List[int], output_path: Path) -> None:
    if not source_path.exists():
        raise FileNotFoundError(f"顺延源文件不存在：{source_path}")
    rows = read_sheet_rows(source_path)
    layout = locate_source_layout(rows)
    header_rows = set(range(1, layout.data_start_index + 1))
    keep_rows = header_rows | {row_no for row_no in kept_source_rows if row_no > layout.data_start_index}
    if len(keep_rows) <= len(header_rows):
        return
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)
    wb = load_workbook(output_path)
    try:
        ws = wb[wb.sheetnames[0]]
        delete_rows = [row_no for row_no in range(layout.data_start_index + 1, ws.max_row + 1) if row_no not in keep_rows]
        for row_no in reversed(delete_rows):
            ws.delete_rows(row_no, 1)
        wb.save(output_path)
    finally:
        close_workbook_safely(wb)


def build_deferred_output_path(output_dir: Path, source_path: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = source_path.stem
    if not stem.endswith("_顺延待处理"):
        stem = f"{stem}_顺延待处理"
    return output_dir / f"{stem}{source_path.suffix}"


def export_deferred_workbooks(output_dir: Path, company_batches: List[CompanyBatch], unassigned_items: List[UnassignedItem]) -> Dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    company_map = {item.company_key: item for item in company_batches}
    deferred_rows: Dict[str, List[int]] = {}
    for item in unassigned_items:
        deferred_rows.setdefault(item.company_key, []).extend(item.source_rows or [])
    exported_paths: Dict[str, Path] = {}
    for company_key, row_numbers in deferred_rows.items():
        company = company_map.get(company_key)
        if company is None:
            continue
        kept_source_rows = sorted({int(row_no) for row_no in row_numbers if int(row_no) > 0})
        if not kept_source_rows:
            continue
        output_path = build_deferred_output_path(output_dir, company.file_path)
        write_deferred_source_workbook(company.file_path, kept_source_rows, output_path)
        if output_path.exists():
            exported_paths[company_key] = output_path
    return exported_paths


def infer_column_width_from_values(values: List[object], cap: int = MAX_EXPORT_COL_WIDTH) -> float:
    width = 8
    for value in values:
        text = "" if value is None else str(value)
        text = text.replace("\n", " ")
        width = max(width, min(len(text) + 2, cap))
    return float(width)


def build_company_process_summary(
    company_batches: List[CompanyBatch],
    assigned_entries: List[AssignedEntry],
    template3_rows_by_row_id: Dict[str, Template3ExportRow],
    unassigned_items: List[UnassignedItem],
) -> List[Tuple[str, str, int, int]]:
    assigned_by_company: Dict[str, int] = {}
    for entry in assigned_entries:
        assigned_by_company[entry.company_key] = assigned_by_company.get(entry.company_key, 0) + 1
    deferred_by_company: Dict[str, int] = {}
    for item in unassigned_items:
        deferred_by_company[item.company_key] = deferred_by_company.get(item.company_key, 0) + 1
    rows: List[Tuple[str, str, int, int]] = []
    for company in company_batches:
        assigned_count = assigned_by_company.get(company.company_key, 0)
        total = company.merged_count
        deferred = deferred_by_company.get(company.company_key, 0)
        export_status = build_export_status(assigned_count, total, deferred)
        rows.append((company.file_path.name, export_status, assigned_count, total))
    return rows


def write_export_workbook(
    output_path: Path,
    company_batches: List[CompanyBatch],
    assigned_entries: List[AssignedEntry],
    bucket_defs: List[BucketDef],
    template2_records: Dict[str, Template2Record],
    template3_rows_by_row_id: Dict[str, Template3ExportRow],
    unassigned_items: List[UnassignedItem],
    exceptions: List[ExceptionItem],
) -> None:
    wb = Workbook()
    try:
        bucket_map = {item.bucket_key: item for item in bucket_defs}
        entry_exact_map, entry_norm_map = build_entry_lookup_maps(assigned_entries)

        ws0 = wb.active
        ws0.title = "批次总览"
        ws0.append(["文件名", "导出时状态", "此次入桶数", "总人数"])
        for file_name, export_status, bucketed_count, total_count in build_company_process_summary(company_batches, assigned_entries, template3_rows_by_row_id, unassigned_items):
            ws0.append([file_name, export_status, bucketed_count, total_count])

        ws1 = wb.create_sheet("标准模板1")
        ws1.append(["桶ID", "公司", "姓名", "职位", "地址", "LinkedIn", "邮箱串", "来源文件"])
        for item in assigned_entries:
            ws1.append([item.bucket_key, item.company, item.name, item.title, item.address, item.linkedin, item.email_str, item.source_file])

        ws2 = wb.create_sheet("模板2输入稿")
        ws2.append(["桶ID", "row_id", "公司", "地址"])
        ordered_entries = sorted(assigned_entries, key=lambda x: (x.target_date_str, x.mode, x.row_id))
        for item in ordered_entries:
            ws2.append([item.bucket_key, item.row_id, item.company, item.address])

        ws3 = wb.create_sheet("发送清单")
        max_email_count = max((len(item.emails) for item in template3_rows_by_row_id.values()), default=1)
        ws3.append(["桶ID", "row_id", "公司", "姓名", "计划发送时间", "状态", *[f"邮箱{i}" for i in range(1, max_email_count + 1)]])
        ordered_rows: List[Tuple[str, str, str, str, List[str]]] = []
        for row_id, record in template2_records.items():
            entry = resolve_entry_by_row_id(row_id, entry_exact_map, entry_norm_map)
            if entry is None:
                continue
            export_row = template3_rows_by_row_id.get(row_id)
            if record.no_valid_time:
                planned = ""
                status = "NO_VALID_TIME"
                emails = split_emails(entry.email_str)
            elif export_row is not None:
                planned = f"{export_row.date_str} {export_row.time_str}"
                status = "已生成"
                emails = export_row.emails
            else:
                planned = f"{record.date_str} {record.time_str}".strip()
                status = "待复核"
                emails = split_emails(entry.email_str)
            ordered_rows.append((entry.bucket_key, row_id, entry.company, entry.name, planned, status, emails))
        ordered_rows.sort(key=lambda x: (x[4], x[0], x[1]))
        for bucket_key, row_id, company, name, planned, status, emails in ordered_rows:
            padded = emails + [""] * (max_email_count - len(emails))
            ws3.append([bucket_key, row_id, company, name, planned, status, *padded])

        ws_delay = wb.create_sheet("顺延清单")
        ws_delay.append(["公司", "姓名", "职位", "LinkedIn", "地址", "邮箱串", "原因", "来源文件"])
        for item in unassigned_items:
            ws_delay.append([item.company, item.name, item.title, item.linkedin, item.address, item.email_str, item.reason, item.source_file])

        ws_err = wb.create_sheet("异常清单")
        ws_err.append(["类型", "row_id", "详情"])
        for item in exceptions:
            ws_err.append([item.category, item.row_id, item.detail])

        for ws in [ws0, ws1, ws2, ws3, ws_delay, ws_err]:
            format_worksheet_dense(ws)
        wb.save(output_path)
    finally:
        close_workbook_safely(wb)


def format_worksheet_dense(ws) -> None:
    default_height = 15
    header_fill = PatternFill(fill_type="solid", fgColor="D9E9FF")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="1F1F1F")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = default_height * 1.5
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = default_height
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    for col_idx in range(1, ws.max_column + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(1, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(col_idx)].width = infer_column_width_from_values(values, MAX_EXPORT_COL_WIDTH)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            left = thin
            right = thin
            top = thin if row_idx == 1 else None
            bottom = thin if row_idx == ws.max_row else None
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    ws.freeze_panes = "A2"


INDEX_HEADERS = [
    "文件名",
    "原始文件路径",
    "文件大小",
    "文件修改时间",
    "导出时状态",
    "最终状态",
    "最近处理日期",
    "完成时间",
    "最近处理桶",
    "此次入桶数",
    "总人数",
    "顺延人数",
    "是否仍有顺延",
    "顺延文件路径",
    "下次最早可处理时间",
    "是否已到可处理时间",
    "对应结果包路径",
    "备注",
]


def parse_index_bool(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "y", "yes", "true", "是", "已到期"}


def get_record_runtime_status(record: IndexRecord) -> str:
    if record.has_deferred and record.deferred_path:
        return "顺延可处理" if record.is_due else "顺延未到期"
    return str(record.final_status or record.export_status or "")


def clear_redo_mark(text: str) -> str:
    raw = str(text or "").strip()
    if not raw:
        return ""
    parts = [part.strip() for part in re.split(r"[，,;；\s]+", raw) if part.strip() and part.strip() != "重做"]
    return "，".join(parts)


def build_export_status(assigned_count: int, total_count: int, deferred_count: int) -> str:
    if assigned_count > 0 and deferred_count > 0:
        return "有顺延"
    if assigned_count > 0:
        return "已入桶"
    if deferred_count > 0:
        return "全部顺延"
    if total_count > 0:
        return "未入桶"
    return "无有效联系人"


def index_record_to_row(record: IndexRecord) -> List[object]:
    return [
        record.file_name,
        record.source_path,
        record.file_size,
        record.file_mtime,
        record.export_status,
        record.final_status,
        record.last_processed_date,
        record.completed_time,
        record.last_bucket,
        record.processed_count,
        record.total_count,
        record.deferred_count,
        "是" if record.has_deferred else "否",
        record.deferred_path,
        record.earliest_reprocess_time,
        "是" if record.is_due else "否",
        record.export_path,
        record.remark,
    ]


def load_index_json(path: Path) -> Dict[str, IndexRecord]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    result: Dict[str, IndexRecord] = {}
    if isinstance(payload, list):
        for item in payload:
            if not isinstance(item, dict):
                continue
            file_name = str(item.get("file_name") or "").strip()
            if not file_name:
                continue
            result[file_name] = IndexRecord(
                file_name=file_name,
                source_path=str(item.get("source_path") or ""),
                file_size=int(item.get("file_size") or 0),
                file_mtime=str(item.get("file_mtime") or ""),
                export_status=str(item.get("export_status") or item.get("status") or ""),
                final_status=str(item.get("final_status") or ""),
                last_processed_date=str(item.get("last_processed_date") or ""),
                completed_time=str(item.get("completed_time") or ""),
                last_bucket=str(item.get("last_bucket") or ""),
                processed_count=int(item.get("processed_count") or 0),
                total_count=int(item.get("total_count") or 0),
                deferred_count=int(item.get("deferred_count") or 0),
                has_deferred=bool(item.get("has_deferred") if item.get("has_deferred") is not None else (item.get("deferred_count") or 0)),
                deferred_path=str(item.get("deferred_path") or ""),
                earliest_reprocess_time=str(item.get("earliest_reprocess_time") or ""),
                is_due=bool(item.get("is_due") or False),
                export_path=str(item.get("export_path") or ""),
                remark=str(item.get("remark") or ""),
            )
    return result


def write_index_json(path: Path, records: Dict[str, IndexRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = []
    for key in sorted(records.keys()):
        item = records[key]
        payload.append({
            "file_name": item.file_name,
            "source_path": item.source_path,
            "file_size": item.file_size,
            "file_mtime": item.file_mtime,
            "export_status": item.export_status,
            "final_status": item.final_status,
            "last_processed_date": item.last_processed_date,
            "completed_time": item.completed_time,
            "last_bucket": item.last_bucket,
            "processed_count": item.processed_count,
            "total_count": item.total_count,
            "deferred_count": item.deferred_count,
            "has_deferred": item.has_deferred,
            "deferred_path": item.deferred_path,
            "earliest_reprocess_time": item.earliest_reprocess_time,
            "is_due": item.is_due,
            "export_path": item.export_path,
            "remark": item.remark,
        })
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_index_xlsx(path: Path) -> Dict[str, IndexRecord]:
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_map: Dict[str, int] = {}
        for cell in ws[1]:
            header_map[str(cell.value or "").strip()] = cell.column

        def get_cell_value(row_idx: int, *header_names: str) -> object:
            for header_name in header_names:
                col_idx = header_map.get(header_name)
                if col_idx:
                    return ws.cell(row=row_idx, column=col_idx).value
            return ""

        result: Dict[str, IndexRecord] = {}
        for row_idx in range(2, ws.max_row + 1):
            file_name = str(get_cell_value(row_idx, "文件名") or "").strip()
            if not file_name:
                continue
            deferred_count_value = int(get_cell_value(row_idx, "顺延人数") or 0)
            result[file_name] = IndexRecord(
                file_name=file_name,
                source_path=str(get_cell_value(row_idx, "原始文件路径") or ""),
                file_size=int(get_cell_value(row_idx, "文件大小") or 0),
                file_mtime=str(get_cell_value(row_idx, "文件修改时间") or ""),
                export_status=str(get_cell_value(row_idx, "导出时状态", "当前状态") or ""),
                final_status=str(get_cell_value(row_idx, "最终状态") or ""),
                last_processed_date=str(get_cell_value(row_idx, "最近处理日期") or ""),
                completed_time=str(get_cell_value(row_idx, "完成时间") or ""),
                last_bucket=str(get_cell_value(row_idx, "最近处理桶") or ""),
                processed_count=int(get_cell_value(row_idx, "此次入桶数", "本次处理人数") or 0),
                total_count=int(get_cell_value(row_idx, "总人数") or 0),
                deferred_count=deferred_count_value,
                has_deferred=parse_index_bool(get_cell_value(row_idx, "是否仍有顺延")) if "是否仍有顺延" in header_map else deferred_count_value > 0,
                deferred_path=str(get_cell_value(row_idx, "顺延文件路径") or ""),
                earliest_reprocess_time=str(get_cell_value(row_idx, "下次最早可处理时间", "最早可再次处理时间") or ""),
                is_due=parse_index_bool(get_cell_value(row_idx, "是否已到可处理时间")),
                export_path=str(get_cell_value(row_idx, "对应结果包路径") or ""),
                remark=str(get_cell_value(row_idx, "备注") or ""),
            )
        return result
    finally:
        close_workbook_safely(wb)


def write_index_xlsx(path: Path, records: Dict[str, IndexRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = "处理索引"
        ws.append(INDEX_HEADERS)
        for key in sorted(records.keys()):
            ws.append(index_record_to_row(records[key]))
        format_worksheet_dense(ws)
        wb.save(path)
    finally:
        close_workbook_safely(wb)


def sync_and_load_index_records(json_path: Path, xlsx_path: Path) -> Dict[str, IndexRecord]:
    xlsx_records = load_index_xlsx(xlsx_path) if xlsx_path.exists() else None
    json_records = load_index_json(json_path) if json_path.exists() else None
    if xlsx_records is not None:
        records = xlsx_records
        write_index_json(json_path, records)
        return records
    if json_records is not None:
        records = json_records
        write_index_xlsx(xlsx_path, records)
        return records
    write_index_json(json_path, {})
    write_index_xlsx(xlsx_path, {})
    return {}


def parse_index_datetime(text: str) -> Optional[datetime]:
    raw = str(text or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=CN_TZ)
        except Exception:
            continue
    return None


def should_force_redo(record: Optional[IndexRecord]) -> bool:
    if record is None:
        return False
    return "重做" in str(record.remark or "")


def refresh_index_due_flags(records: Dict[str, IndexRecord], now_dt: datetime) -> Dict[str, IndexRecord]:
    for item in records.values():
        item.has_deferred = bool(item.has_deferred or item.deferred_count > 0)
        if item.has_deferred and item.deferred_path:
            earliest = parse_index_datetime(item.earliest_reprocess_time)
            item.is_due = True if earliest is None else now_dt >= earliest
        else:
            item.is_due = False
        if not item.export_status:
            item.export_status = build_export_status(item.processed_count, item.total_count, item.deferred_count)
    return records


def has_file_changed(path: Path, record: Optional[IndexRecord]) -> bool:
    if record is None or not path.exists():
        return False
    stat = path.stat()
    current_size = int(stat.st_size)
    current_mtime = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    return current_size != int(record.file_size or 0) or current_mtime != str(record.file_mtime or "")


def sync_company_batch_file_paths(
    company_batches: List[CompanyBatch],
    current_source_paths: List[Path],
    index_records: Dict[str, IndexRecord],
) -> None:
    live_by_name: Dict[str, Path] = {}
    for path in current_source_paths:
        try:
            if path.exists():
                live_by_name[path.name] = path
        except Exception:
            continue
    for company in company_batches:
        try:
            if company.file_path.exists():
                live_by_name[company.file_path.name] = company.file_path
        except Exception:
            continue
    for company in company_batches:
        try:
            if company.file_path.exists():
                continue
        except Exception:
            pass
        candidate = live_by_name.get(company.file_path.name)
        if candidate is None:
            record = index_records.get(company.file_path.name)
            if record is not None:
                for raw_path in [record.source_path, record.deferred_path]:
                    raw_text = str(raw_path or "").strip()
                    if not raw_text:
                        continue
                    path_obj = Path(raw_text)
                    if path_obj.exists():
                        candidate = path_obj
                        break
        if candidate is not None:
            company.file_path = candidate


def apply_archive_map_to_company_batches(company_batches: List[CompanyBatch], archive_map: Dict[str, Path]) -> None:
    for company in company_batches:
        source_keys = []
        try:
            source_keys.append(str(company.file_path.resolve()))
        except Exception:
            pass
        source_keys.append(company.company_key)
        for source_key in source_keys:
            target_path = archive_map.get(source_key)
            if target_path is not None:
                company.file_path = target_path
                break


def build_export_error_message(step_name: str, exc: Exception, related_path: Optional[Path] = None) -> str:
    lines = [f"{step_name}失败。"]
    if isinstance(exc, PermissionError):
        lines.append("文件正被其他程序占用，程序没有写入权限。")
        lines.append("请先关闭 Excel / WPS / 文件预览窗口，然后重试。")
    elif isinstance(exc, FileNotFoundError):
        lines.append("程序要操作的文件不存在。")
        lines.append("常见原因：上一次导出后文件已被移入日期文件夹，但当前批次仍引用旧路径。")
        lines.append("处理方式：重新导入当前公司文件后，再重新导出。")
    else:
        lines.append("程序在该步骤抛出了异常。")
    if related_path is not None:
        lines.append(f"相关文件：{related_path}")
    lines.append(f"系统信息：{exc}")
    return "\n".join(lines)


def collect_startup_candidates(source_dir: Path, index_records: Dict[str, IndexRecord]) -> Tuple[List[StartupCandidate], List[str]]:
    source_dir.mkdir(parents=True, exist_ok=True)
    deferred_dir = source_dir / "顺延待处理"
    root_files = [path for path in sorted(source_dir.iterdir()) if path.is_file() and DropFrame._is_excel(str(path)) and not path.name.startswith("~$")]
    candidates: List[StartupCandidate] = []
    skipped_messages: List[str] = []
    for path in root_files:
        record = index_records.get(path.name)
        if record is None or should_force_redo(record):
            label = "新文件" if record is None else "备注重做"
            candidates.append(StartupCandidate(file_path=path, file_name=path.name, label=label, checked=True, is_deferred=False))
            continue
        if has_file_changed(path, record):
            candidates.append(StartupCandidate(file_path=path, file_name=path.name, label="已检测到文件变更，按重做处理", checked=True, is_deferred=False))
            continue
        skipped_messages.append(f"{path.name}|{get_record_runtime_status(record)}")
    for file_name, record in sorted(index_records.items(), key=lambda x: x[0]):
        if not record.deferred_path or not record.is_due:
            continue
        deferred_path = Path(record.deferred_path)
        if not deferred_path.exists() and deferred_dir.exists():
            fallback = deferred_dir / Path(record.deferred_path).name
            deferred_path = fallback
        if deferred_path.exists():
            label = f"{get_record_runtime_status(record)} | 最早可处理 {record.earliest_reprocess_time or '-'} | 顺延{record.deferred_count}人"
            candidates.append(StartupCandidate(file_path=deferred_path, file_name=deferred_path.name, label=label, checked=True, is_deferred=True))
    return candidates, skipped_messages


def archive_processed_source_files(source_paths: List[Path], source_dir: Optional[Path], date_str: str) -> Dict[str, Path]:
    archive_map: Dict[str, Path] = {}
    for source_path in source_paths:
        if not source_path.exists():
            continue
        base_dir = source_dir if source_dir is not None else source_path.parent
        archive_dir = base_dir / date_str
        archive_dir.mkdir(parents=True, exist_ok=True)
        try:
            if source_path.parent.resolve() == archive_dir.resolve():
                archive_map[str(source_path.resolve())] = source_path
                continue
        except Exception:
            pass
        target_path = build_non_overwriting_path(archive_dir, source_path.stem, source_path.suffix)
        shutil.move(str(source_path), str(target_path))
        archive_map[str(source_path.resolve())] = target_path
    return archive_map


def next_workday_date(day_value) -> object:
    current = day_value + timedelta(days=1)
    while current.weekday() >= 5:
        current += timedelta(days=1)
    return current


def compute_company_earliest_reprocess_time(
    company_key: str,
    unassigned_items: List[UnassignedItem],
    bucket_defs: List[BucketDef],
) -> str:
    company_items = [item for item in unassigned_items if item.company_key == company_key]
    if not company_items:
        return ""
    if bucket_defs:
        last_bucket_date = max(parse_cn_date(item.target_date_str).date() for item in bucket_defs)
    else:
        last_bucket_date = get_current_cn_datetime().date()
    candidate_date = last_bucket_date + timedelta(days=7)
    candidate_dt = datetime.combine(candidate_date, datetime.min.time()).replace(tzinfo=CN_TZ)
    return candidate_dt.strftime("%Y-%m-%d %H:%M:%S")


def update_index_after_export(
    existing_records: Dict[str, IndexRecord],
    company_batches: List[CompanyBatch],
    assigned_entries: List[AssignedEntry],
    template3_rows_by_row_id: Dict[str, Template3ExportRow],
    unassigned_items: List[UnassignedItem],
    export_path: Path,
    archive_map: Dict[str, Path],
    deferred_map: Dict[str, Path],
    bucket_defs: List[BucketDef],
) -> Dict[str, IndexRecord]:
    result = copy.deepcopy(existing_records)
    processed_by_company: Dict[str, int] = {}
    last_bucket_by_company: Dict[str, str] = {}
    for entry in assigned_entries:
        processed_by_company[entry.company_key] = processed_by_company.get(entry.company_key, 0) + 1
        last_bucket_by_company[entry.company_key] = entry.bucket_key
    deferred_by_company: Dict[str, int] = {}
    for item in unassigned_items:
        deferred_by_company[item.company_key] = deferred_by_company.get(item.company_key, 0) + 1

    current_dt = get_current_cn_datetime().strftime("%Y-%m-%d %H:%M:%S")
    current_date = get_current_cn_datetime().strftime("%Y-%m-%d")
    for company in company_batches:
        file_name = company.file_path.name
        archived_path = archive_map.get(str(company.file_path.resolve()))
        size_value = 0
        mtime_value = ""
        if archived_path is not None and archived_path.exists():
            stat = archived_path.stat()
            size_value = int(stat.st_size)
            mtime_value = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        deferred_count = deferred_by_company.get(company.company_key, 0)
        processed_count = processed_by_company.get(company.company_key, 0)
        earliest = compute_company_earliest_reprocess_time(company.company_key, unassigned_items, bucket_defs)
        deferred_path = str(deferred_map.get(company.company_key) or "")
        previous = result.get(file_name)
        export_status = build_export_status(processed_count, company.merged_count, deferred_count)
        record = IndexRecord(
            file_name=file_name,
            source_path=str(archived_path or company.file_path),
            file_size=size_value,
            file_mtime=mtime_value,
            export_status=export_status,
            final_status="",
            last_processed_date=current_date,
            completed_time="",
            last_bucket=last_bucket_by_company.get(company.company_key, ""),
            processed_count=processed_count,
            total_count=company.merged_count,
            deferred_count=deferred_count,
            has_deferred=deferred_count > 0,
            deferred_path=deferred_path,
            earliest_reprocess_time=earliest,
            is_due=False,
            export_path=str(export_path),
            remark=clear_redo_mark(previous.remark if previous is not None else ""),
        )
        result[file_name] = record
    return refresh_index_due_flags(result, get_current_cn_datetime())

def main() -> None:
    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
